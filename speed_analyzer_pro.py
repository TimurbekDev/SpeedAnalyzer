#!/usr/bin/env python3
"""
Speed Analyzer Pro  —  Birlashgan Tezlik Analizatori
=====================================================
camera_calibrator.py + vehicle_detector.py + speed_analyzer_v2.py

Xususiyatlar:
  ✓ YOLO (yolov8n.pt) yoki MOG2 background subtraction
  ✓ ORB video stabilizatsiya (kamera titroq kompensatsiyasi)
  ✓ Perspektiv-to'g'ri tezlik (kamera balandligi + tilt + fokus)
  ✓ ROI (qiziqish hududi) tanlash
  ✓ Kamera kalibrlash oynasi (vanishing point / qo'lda / 2 nuqta)
  ✓ Ko'p oqimli arxitektura — UI muzlamaydi
  ✓ Excel eksport + avtomatik saqlash
  ✓ Yo'l izi, ID yorliq, screenshot

Ishga tushirish:
    pip install opencv-python numpy pillow ultralytics openpyxl
    python speed_analyzer_pro.py
"""

import cv2
import numpy as np
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk
import threading
import queue
import time
import os
import json
from datetime import datetime
from collections import deque

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ═══════════════════════════════════════════════════════════════════════════════
#  Constants
# ═══════════════════════════════════════════════════════════════════════════════
VEHICLE_CLASSES = {2, 5, 7}       # COCO: car, bus, truck
CONF_DEFAULT    = 0.35
FRAME_SKIP      = 2               # YOLO every N frames
CAP_Q_SIZE      = 4
RES_Q_SIZE      = 6

BG_C  = "#0a0c12"
SB_C  = "#0d1018"
ACC_C = "#f59e0b"

SPEED_COLORS = [(60, (0, 220, 80)), (110, (0, 200, 255)), (999, (0, 60, 255))]


# ═══════════════════════════════════════════════════════════════════════════════
#  YOLO loader
# ═══════════════════════════════════════════════════════════════════════════════
def load_yolo(path=None):
    try:
        from ultralytics import YOLO
        import torch
        device = "cuda" if torch.cuda.is_available() else "cpu"
        candidates = ([path] if path else []) + [
            "yolov8n.pt", "yolov8s.pt", "yolov9c.pt", "best.pt", "yolov10n.pt"]
        for c in candidates:
            if c and os.path.exists(c):
                try:
                    m = YOLO(c)
                    if device == "cuda":
                        m.to("cuda")
                    print(f"[YOLO] Loaded {c} on {device}")
                    return m, os.path.basename(c)
                except Exception as e:
                    print(f"[YOLO] {c}: {e}")
        m = YOLO("yolov8n.pt")
        if device == "cuda":
            m.to("cuda")
        return m, "yolov8n.pt"
    except Exception as e:
        print(f"[YOLO] Not available: {e}")
        return None, None


# ═══════════════════════════════════════════════════════════════════════════════
#  Enhancer
# ═══════════════════════════════════════════════════════════════════════════════
class Enhancer:
    def __init__(self):
        self.clahe   = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8, 8))
        self.enabled = True

    def process(self, frame: np.ndarray) -> np.ndarray:
        if not self.enabled or np.mean(frame) > 90:
            return frame
        lab = cv2.cvtColor(frame, cv2.COLOR_BGR2LAB)
        l, a, b = cv2.split(lab)
        return cv2.cvtColor(cv2.merge([self.clahe.apply(l), a, b]), cv2.COLOR_LAB2BGR)


# ═══════════════════════════════════════════════════════════════════════════════
#  VideoStabilizer
# ═══════════════════════════════════════════════════════════════════════════════
class VideoStabilizer:
    def __init__(self):
        self.enabled        = True
        self.stabilize_view = True
        self.orb = cv2.ORB_create(nfeatures=500, scoreType=cv2.ORB_HARRIS_SCORE)
        self.bf  = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=False)
        self.prev_gray = self.prev_kp = self.prev_des = None
        self.shift_history  = deque(maxlen=30)
        self.dx = self.dy   = 0.0
        self.confidence     = 0.0
        self.total_frames   = 0
        self.stable_frames  = 0

    def reset(self):
        self.prev_gray = self.prev_kp = self.prev_des = None
        self.shift_history.clear()
        self.dx = self.dy = 0.0
        self.total_frames = self.stable_frames = 0

    def process(self, frame: np.ndarray):
        """Returns (stabilized_frame, dx, dy, confidence)."""
        self.total_frames += 1
        if not self.enabled:
            return frame, 0.0, 0.0, 0.0

        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        kp, des = self.orb.detectAndCompute(gray, None)

        if self.prev_gray is None or des is None or self.prev_des is None:
            self.prev_gray, self.prev_kp, self.prev_des = gray, kp, des
            return frame, 0.0, 0.0, 0.0

        try:
            matches = self.bf.knnMatch(self.prev_des, des, k=2)
        except Exception:
            self.prev_gray, self.prev_kp, self.prev_des = gray, kp, des
            return frame, 0.0, 0.0, 0.0

        good = []
        for pair in matches:
            if len(pair) == 2:
                m, n = pair
                if m.distance < 0.75 * n.distance:
                    good.append(m)

        self.confidence = len(good) / max(len(matches), 1)

        if len(good) < 10:
            self.prev_gray, self.prev_kp, self.prev_des = gray, kp, des
            self.dx = self.dy = 0.0
            return frame, 0.0, 0.0, 0.0

        src = np.float32([self.prev_kp[m.queryIdx].pt for m in good]).reshape(-1, 1, 2)
        dst = np.float32([kp[m.trainIdx].pt for m in good]).reshape(-1, 1, 2)
        M, _ = cv2.estimateAffinePartial2D(src, dst, method=cv2.RANSAC,
                                           ransacReprojThreshold=3.0)

        if M is None:
            self.prev_gray, self.prev_kp, self.prev_des = gray, kp, des
            self.dx = self.dy = 0.0
            return frame, 0.0, 0.0, 0.0

        self.dx, self.dy = float(M[0, 2]), float(M[1, 2])
        self.shift_history.append((self.dx, self.dy))
        if np.hypot(self.dx, self.dy) < 1.0:
            self.stable_frames += 1

        result = frame
        if self.stabilize_view and np.hypot(self.dx, self.dy) > 0.5:
            h, w = frame.shape[:2]
            sx = np.mean([s[0] for s in self.shift_history]) if len(self.shift_history) >= 3 else self.dx
            sy = np.mean([s[1] for s in self.shift_history]) if len(self.shift_history) >= 3 else self.dy
            result = cv2.warpAffine(frame, np.float32([[1, 0, -sx], [0, 1, -sy]]), (w, h),
                                    borderMode=cv2.BORDER_REPLICATE)

        self.prev_gray, self.prev_kp, self.prev_des = gray, kp, des
        return result, self.dx, self.dy, self.confidence

    @property
    def status_text(self):
        if not self.enabled:
            return "O'chiq"
        pct = (self.stable_frames / max(self.total_frames, 1)) * 100
        return f"dx={self.dx:.1f} dy={self.dy:.1f} ({pct:.0f}% stabil)"


# ═══════════════════════════════════════════════════════════════════════════════
#  CameraModel
# ═══════════════════════════════════════════════════════════════════════════════
class CameraModel:
    def __init__(self):
        self.enabled      = False
        self.H            = 6.0
        self.tilt         = 15.0
        self.pan          = 0.0
        self.f            = 800.0
        self.cx           = 640.0
        self.cy           = 360.0
        self.px_per_m     = 20.0
        self.speed_factor = 1.0   # user-adjustable multiplier for final output

    def setup(self, H, tilt, pan, f, fw, fh):
        self.H, self.tilt, self.pan, self.f = H, tilt, pan, f
        self.cx, self.cy = fw / 2.0, fh / 2.0
        self.enabled = True

    def load_json(self, path):
        with open(path) as fp:
            d = json.load(fp)
        self.H     = d.get("camera_height_m", 6.0)
        self.tilt  = d.get("camera_tilt_deg", d.get("camera_pitch_deg", 15.0))
        self.pan   = d.get("camera_pan_deg", 0.0)
        self.f     = d.get("focal_length_px", 800.0)
        self.cx    = d.get("frame_width",  1280) / 2.0
        self.cy    = d.get("frame_height",  720) / 2.0
        self.enabled = True

    def pixel_speed_to_kmh(self, vx_px_s: float, vy_px_s: float, ref_v: float) -> float:
        """Convert pixel-space velocity components (px/s) to km/h.

        Uses direction-aware scales:
          horizontal component → lateral scale  = Z/f
          vertical component   → depth scale    = Z²/(f·H)
        where Z is the camera-frame depth to the ground point at row ref_v.
        This gives correct speed regardless of whether the vehicle crosses
        laterally, approaches, or moves diagonally.
        """
        sf = max(0.01, self.speed_factor)
        if self.enabled:
            tilt_r  = np.radians(self.tilt)
            alpha_v = np.arctan((self.cy - ref_v) / self.f)
            beta    = tilt_r - alpha_v
            if beta < 0.015:
                return 0.0
            # Camera-frame depth Z = H·cos(alpha_v)/sin(beta)
            Z = self.H * np.cos(alpha_v) / np.sin(beta)
            if Z < 0.1:
                return 0.0
            m_per_px_x = Z / self.f                       # lateral scale
            m_per_px_y = (Z * Z) / (self.f * self.H)      # depth scale
            speed_m_s  = np.sqrt((vx_px_s * m_per_px_x) ** 2 +
                                  (vy_px_s * m_per_px_y) ** 2)
            return speed_m_s * 3.6 * sf
        px_per_s = np.sqrt(vx_px_s ** 2 + vy_px_s ** 2)
        return px_per_s / self.px_per_m * 3.6 * sf


# ═══════════════════════════════════════════════════════════════════════════════
#  SpeedEstimator — per-track stable velocity estimation
# ═══════════════════════════════════════════════════════════════════════════════
class SpeedEstimator:
    """
    Stable, real-time speed estimator for a single track.

    Pipeline per frame:
      raw centroid (cx,cy)
        → alpha-beta filter        (removes positional jitter)
        → rolling buffer (≤20 pts)
        → windowed weighted polyfit (velocity in px/s)
        → CameraModel conversion    (px/s → m/s → km/h)
        → median spike rejection    (kills outlier bursts)
        → EWMA output smoothing     (final stable km/h)

    ID re-association:  call inherit(other) to copy history from a
    recently-dead track so speed continuity is preserved across ID switches.
    """

    WINDOW_S   = 0.9    # regression window (seconds)
    AB_ALPHA   = 0.35   # alpha-beta: position correction gain
    AB_BETA    = 0.15   # alpha-beta: velocity correction gain
    JUMP_THR   = 80     # px jump threshold — reduce gain above this
    EWMA_A     = 0.25   # output smoothing   (lower → smoother but laggier)
    SPIKE_K    = 2.8    # reject raw estimate if > K × recent_median
    MIN_PTS    = 3      # minimum buffer points for reliable fit
    MIN_DT     = 0.06   # minimum time span in window (s)
    MAX_KMH    = 200    # hard physical cap

    def __init__(self, fps: float):
        self.fps         = fps
        self._buf        = deque(maxlen=20)   # (fid, t, sx, sy)  smoothed positions
        self._speed_buf  = deque(maxlen=8)    # recent raw km/h estimates (spike window)
        self.speed_kmh   = 0.0               # final output
        # Alpha-beta state
        self._sx = self._sy = None           # smoothed position
        self._vx = self._vy = 0.0            # estimated velocity (px/frame)

    # ── Public API ────────────────────────────────────────────────────────────
    def push(self, fid: int, cx: float, cy: float) -> None:
        """Feed a new raw centroid. Runs alpha-beta filter, stores smoothed point."""
        t = fid / self.fps
        if self._sx is None:
            self._sx, self._sy = cx, cy
            self._buf.append((fid, t, cx, cy))
            return
        # Predict step
        px_pred = self._sx + self._vx
        py_pred = self._sy + self._vy
        rx, ry  = cx - px_pred, cy - py_pred
        # Reduce gain on implausibly large jumps (bbox jitter / occlusion)
        jump  = np.hypot(rx, ry)
        alpha = self.AB_ALPHA if jump < self.JUMP_THR else self.AB_ALPHA * 0.25
        beta  = self.AB_BETA  if jump < self.JUMP_THR else 0.0
        # Update step
        self._sx  = px_pred + alpha * rx
        self._sy  = py_pred + alpha * ry
        self._vx += beta * rx
        self._vy += beta * ry
        self._buf.append((fid, t, self._sx, self._sy))

    def estimate(self, cam: CameraModel, ref_row: float) -> float:
        """Compute current speed from smoothed buffer. Returns stable km/h."""
        if len(self._buf) < self.MIN_PTS:
            return self.speed_kmh

        t_now = self._buf[-1][1]
        pts   = [(t, sx, sy) for _, t, sx, sy in self._buf
                 if t_now - t <= self.WINDOW_S]
        if len(pts) < self.MIN_PTS:
            return self.speed_kmh

        times = np.array([p[0] for p in pts])
        xs    = np.array([p[1] for p in pts])
        ys    = np.array([p[2] for p in pts])

        dt = times[-1] - times[0]
        if dt < self.MIN_DT:
            return self.speed_kmh

        # Exponentially-weighted polyfit → velocity components in px/s
        w    = np.exp(np.linspace(-2.0, 0.0, len(times)))
        vx_s = np.polyfit(times, xs, 1, w=w)[0]
        vy_s = np.polyfit(times, ys, 1, w=w)[0]

        # Convert pixel-space velocity → km/h using direction-aware camera model
        raw = cam.pixel_speed_to_kmh(vx_s, vy_s, ref_row)
        if not (0.0 < raw < self.MAX_KMH):
            return self.speed_kmh

        # ── Spike rejection ───────────────────────────────────────────────
        if len(self._speed_buf) >= 4:
            med = float(np.median(list(self._speed_buf)))
            if med > 1.0 and raw > med * self.SPIKE_K:
                return self.speed_kmh   # discard outlier burst

        self._speed_buf.append(raw)

        # ── EWMA output smoothing ─────────────────────────────────────────
        self.speed_kmh = (raw if self.speed_kmh < 0.5
                          else self.EWMA_A * raw + (1.0 - self.EWMA_A) * self.speed_kmh)
        return self.speed_kmh

    def inherit(self, other: "SpeedEstimator") -> None:
        """
        Copy position history + speed state from another estimator.
        Called when a new track is associated with a recently-dead one
        (ID switch recovery) — preserves speed continuity.
        """
        self._buf       = deque(other._buf,       maxlen=other._buf.maxlen)
        self._speed_buf = deque(other._speed_buf, maxlen=other._speed_buf.maxlen)
        self.speed_kmh  = other.speed_kmh
        self._sx, self._sy = other._sx, other._sy
        self._vx, self._vy = other._vx, other._vy

    def reset(self) -> None:
        self._buf.clear(); self._speed_buf.clear()
        self.speed_kmh = 0.0
        self._sx = self._sy = None
        self._vx = self._vy = 0.0


# ═══════════════════════════════════════════════════════════════════════════════
#  Background subtractor (no-YOLO fallback)
# ═══════════════════════════════════════════════════════════════════════════════
class BGDetector:
    def __init__(self, min_a=1500, max_a=150000):
        self.min_a = min_a
        self.max_a = max_a
        self.bg       = cv2.createBackgroundSubtractorMOG2(
            history=500, varThreshold=50, detectShadows=True)
        self.k_open   = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (5, 5))
        self.k_close  = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (13, 13))
        self.k_dilate = cv2.getStructuringElement(cv2.MORPH_RECT,    (11, 11))

    def detect(self, frame: np.ndarray, roi_mask=None) -> list:
        gray = cv2.GaussianBlur(cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY), (7, 7), 0)
        fg   = self.bg.apply(gray, learningRate=0.003)
        fg   = np.where(fg == 255, 255, 0).astype(np.uint8)
        fg   = cv2.morphologyEx(fg, cv2.MORPH_OPEN,  self.k_open)
        fg   = cv2.morphologyEx(fg, cv2.MORPH_CLOSE, self.k_close)
        fg   = cv2.dilate(fg, self.k_dilate, iterations=2)
        if roi_mask is not None:
            fg = cv2.bitwise_and(fg, roi_mask)
        cnts, _ = cv2.findContours(fg, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        boxes = []
        for c in cnts:
            x, y, w, h = cv2.boundingRect(c)
            a = w * h
            if not (self.min_a < a < self.max_a):
                continue
            if not (0.2 < w / max(h, 1) < 5.5):
                continue
            boxes.append((x, y, w, h))
        return boxes

    def update_area(self, mn, mx):
        self.min_a, self.max_a = mn, mx


# ═══════════════════════════════════════════════════════════════════════════════
#  YOLO Detector wrapper
# ═══════════════════════════════════════════════════════════════════════════════
class YOLODetector:
    def __init__(self, model, conf=CONF_DEFAULT):
        self.model  = model
        self.conf   = conf
        try:
            import torch
            self.device = "cuda" if torch.cuda.is_available() else "cpu"
        except Exception:
            self.device = "cpu"

    def detect(self, frame: np.ndarray, roi_mask=None) -> list:
        if self.model is None:
            return []
        res = self.model(frame, verbose=False, conf=self.conf,
                         classes=list(VEHICLE_CLASSES),
                         half=(self.device == "cuda"))[0]
        boxes = []
        for b in res.boxes:
            x1, y1, x2, y2 = map(int, b.xyxy[0])
            w, h = x2 - x1, y2 - y1
            if roi_mask is not None:
                vm   = np.zeros(roi_mask.shape, np.uint8)
                vy1_ = max(0, y1); vy2_ = min(roi_mask.shape[0], y2)
                vx1_ = max(0, x1); vx2_ = min(roi_mask.shape[1], x2)
                vm[vy1_:vy2_, vx1_:vx2_] = 255
                overlap = np.count_nonzero(cv2.bitwise_and(vm, roi_mask)) / max(w * h, 1)
                if overlap < 0.3:
                    continue
            boxes.append((x1, y1, w, h))
        return boxes


# ═══════════════════════════════════════════════════════════════════════════════
#  Track
# ═══════════════════════════════════════════════════════════════════════════════
class Track:
    _next = 1

    def __init__(self, box: tuple, fid: int, fps: float):
        self.id            = Track._next; Track._next += 1
        self.bbox          = box
        self.cx            = box[0] + box[2] / 2.0
        self.cy            = box[1] + box[3] / 2.0
        self.fps           = fps
        self.age           = 0
        self.hits          = 1
        self.speed_kmh     = 0.0
        self.speed_samples: list = []
        self.first_fid     = fid
        self.last_fid      = fid
        self.history       = deque(maxlen=150)
        self.history.append((fid, self.cx, self.cy))
        # Speed estimator (alpha-beta + EWMA + spike rejection)
        self._est = SpeedEstimator(fps)
        self._est.push(fid, self.cx, self.cy)
        # Line crossing state
        self.prev_cx   = self.cx
        self.prev_cy   = self.cy
        self.line_side = None   # last known side of counting line (+1 / -1)
        self.crossed   = False  # already crossed (counted)
        self.cross_dir = None   # "IN" or "OUT"
        # Unique color
        np.random.seed(self.id * 37 + 7)
        h = (self.id * 47) % 180
        self.color = tuple(int(c) for c in
                           cv2.cvtColor(np.uint8([[[h, 200, 230]]]),
                                        cv2.COLOR_HSV2BGR)[0][0])

    def predict(self):
        if len(self.history) < 2:
            return self.cx, self.cy
        vx = self.history[-1][1] - self.history[-2][1]
        vy = self.history[-1][2] - self.history[-2][2]
        return self.history[-1][1] + vx, self.history[-1][2] + vy

    def update(self, box: tuple, fid: int):
        self.prev_cx  = self.cx
        self.prev_cy  = self.cy
        self.bbox     = box
        self.cx       = box[0] + box[2] / 2.0
        self.cy       = box[1] + box[3] / 2.0
        self.age      = 0
        self.hits    += 1
        self.last_fid = fid
        self.history.append((fid, self.cx, self.cy))

    def calc_speed(self, cam: CameraModel):
        ref_row = self.cy + self.bbox[3] / 2.0
        self._est.push(self.last_fid, self.cx, self.cy)
        kmh = self._est.estimate(cam, ref_row)
        self.speed_kmh = kmh
        if kmh > 0:
            self.speed_samples.append(kmh)

    @property
    def trail(self):
        return [(int(h[1]), int(h[2])) for h in self.history]

    @property
    def duration(self):
        if len(self.history) < 2:
            return 0.0
        return (self.history[-1][0] - self.history[0][0]) / self.fps


# ═══════════════════════════════════════════════════════════════════════════════
#  Tracker
# ═══════════════════════════════════════════════════════════════════════════════
class Tracker:
    def __init__(self, fps: float, cam: CameraModel,
                 max_age=20, min_hits=3, iou_thr=0.2, dist_thr=120):
        self.fps      = fps
        self.cam      = cam
        self.max_age  = max_age
        self.min_hits = min_hits
        self.iou_thr  = iou_thr
        self.dist_thr = dist_thr
        self.tracks:   list[Track] = []
        self.finished: list[Track] = []
        # Counted-track state — persists for full session
        self.counted_ids: set = set()          # IDs already crossed
        self._ghost_buf:  deque = deque(maxlen=60)  # (cx, cy, fid) of retired crossers
        # Dead-estimator pool for ID-switch recovery
        self._dead_estimators: deque = deque(maxlen=20)  # (cx, cy, fid, SpeedEstimator)

    def reset(self):
        self.tracks      = []
        self.finished    = []
        Track._next      = 1
        self.counted_ids.clear()
        self._ghost_buf.clear()
        self._dead_estimators.clear()

    # ── Ghost guard: mark new tracks near a recently-counted position ─────────
    def _is_ghost(self, cx: float, cy: float, fid: int,
                  radius: float = 90, max_frames: int = 45) -> bool:
        """Return True if (cx,cy) is within `radius` px of a recently counted position."""
        return any(
            np.hypot(cx - px, cy - py) < radius and (fid - pf) < max_frames
            for px, py, pf in self._ghost_buf
        )

    def _make_track(self, box: tuple, fid: int) -> Track:
        """Create a new Track, pre-marking it as crossed if it's a ghost re-detection.
        Also inherits SpeedEstimator state from a nearby recently-dead track (ID switch recovery)."""
        t = Track(box, fid, self.fps)
        x, y, w, h = box
        cx, cy = x + w / 2.0, y + h / 2.0
        if self._is_ghost(cx, cy, fid):
            t.crossed = True   # inherit counted state — skip line checks
            return t
        # Try to inherit estimator from a recently-dead nearby track
        best_dist, best_est = 90.0, None
        for (dcx, dcy, dfid, dest) in self._dead_estimators:
            if fid - dfid > 45:
                continue
            d = np.hypot(cx - dcx, cy - dcy)
            if d < best_dist:
                best_dist, best_est = d, dest
        if best_est is not None:
            t._est.inherit(best_est)
            t.speed_kmh = best_est.speed_kmh
        return t

    def update(self, boxes: list, fid: int) -> list:
        for t in self.tracks:
            t.age += 1

        if boxes and self.tracks:
            nt, nd = len(self.tracks), len(boxes)
            cost   = np.full((nt, nd), 1e9)
            for i, trk in enumerate(self.tracks):
                px, py = trk.predict()
                tx, ty, tw, th = trk.bbox
                for j, (x, y, w, h) in enumerate(boxes):
                    iou  = self._iou((tx, ty, tx+tw, ty+th), (x, y, x+w, y+h))
                    dist = np.hypot(px-(x+w/2), py-(y+h/2))
                    if iou > self.iou_thr or dist < self.dist_thr:
                        cost[i, j] = (1 - iou) * 80 + dist
            mt, md = set(), set()
            for idx in np.argsort(cost, axis=None):
                i, j = divmod(int(idx), nd)
                if cost[i, j] >= 1e8:
                    break
                if i not in mt and j not in md:
                    self.tracks[i].update(boxes[j], fid)
                    mt.add(i); md.add(j)
            for j in range(nd):
                if j not in md:
                    self.tracks.append(self._make_track(boxes[j], fid))
        elif boxes:
            for b in boxes:
                self.tracks.append(self._make_track(b, fid))

        # ── Retire tracks: age-out OR immediately after crossing ─────────────
        alive = []
        for t in self.tracks:
            if t.crossed:
                # Logical removal — crossed tracks leave the active pool at once.
                # This prevents the matcher from binding future detections of the
                # same physical vehicle to this ID again.
                self.counted_ids.add(t.id)
                self._ghost_buf.append((t.cx, t.cy, fid))   # poison zone
                if t.hits >= self.min_hits and t.speed_samples:
                    self.finished.append(t)
                # do NOT add to alive → removed immediately
            elif t.age > self.max_age:
                # Stash estimator so a new track can inherit speed continuity
                self._dead_estimators.append((t.cx, t.cy, fid, t._est))
                if t.hits >= self.min_hits and t.speed_samples:
                    self.finished.append(t)
            else:
                alive.append(t)
        self.tracks = alive
        return [t for t in self.tracks if t.hits >= self.min_hits]

    @staticmethod
    def _iou(b1, b2):
        xi1 = max(b1[0], b2[0]); yi1 = max(b1[1], b2[1])
        xi2 = min(b1[2], b2[2]); yi2 = min(b1[3], b2[3])
        inter = max(0, xi2-xi1) * max(0, yi2-yi1)
        a1 = (b1[2]-b1[0]) * (b1[3]-b1[1])
        a2 = (b2[2]-b2[0]) * (b2[3]-b2[1])
        return inter / max(a1 + a2 - inter, 1e-6)


# ═══════════════════════════════════════════════════════════════════════════════
#  Worker Thread  (capture-queue → detect+track → result-queue)
# ═══════════════════════════════════════════════════════════════════════════════
class Worker(threading.Thread):
    def __init__(self, yolo_det, bg_det, enhancer, tracker, in_q, out_q,
                 cross_line=None):
        super().__init__(daemon=True)
        self.yolo  = yolo_det
        self.bg    = bg_det
        self.enh   = enhancer
        self.trk   = tracker
        self.in_q  = in_q
        self.out_q = out_q
        self._stop = threading.Event()
        self._fid  = 0
        self._last_boxes: list = []
        self._roi_mask:   np.ndarray | None = None
        self._cross_line  = cross_line   # CrossLine ref (set by main thread)
        self._lock = threading.Lock()

    def set_roi(self, mask):
        with self._lock:
            self._roi_mask = mask

    def set_cross_line(self, cl):
        with self._lock:
            self._cross_line = cl

    @staticmethod
    def _in_roi(cx: float, cy: float, mask) -> bool:
        """True ONLY when a ROI mask is defined AND the point is inside it.
        No ROI → no speed estimation (speed requires a calibrated zone)."""
        if mask is None:
            return False
        xi, yi = int(cx), int(cy)
        h, w = mask.shape[:2]
        return 0 <= yi < h and 0 <= xi < w and mask[yi, xi] > 0

    def run(self):
        while not self._stop.is_set():
            try:
                item = self.in_q.get(timeout=0.05)
            except queue.Empty:
                continue
            if item is None:
                self.out_q.put(None)
                break

            fid, frame = item
            enhanced   = self.enh.process(frame)
            with self._lock:
                roi = self._roi_mask
                cl  = self._cross_line

            if self._fid % max(1, FRAME_SKIP) == 0:
                if self.yolo and self.yolo.model:
                    # Detect on full frame — ROI must NOT limit tracking or counting.
                    # Speed filtering is applied later, per centroid position.
                    self._last_boxes = self.yolo.detect(enhanced, None)
                else:
                    # BG subtractor is noisy on the full frame; keep ROI mask here.
                    self._last_boxes = self.bg.detect(enhanced, roi)
            self._fid += 1

            tracks   = self.trk.update(self._last_boxes, fid)
            finished = self.trk.finished[:]
            self.trk.finished.clear()

            # ── Speed calculation: only when centroid is inside ROI ───────────
            # Avoids wasting CPU on vehicles outside the measurement zone and
            # prevents stale/wrong speeds from tracks that left the ROI.
            cam = self.trk.cam
            for t in tracks:
                if self._in_roi(t.cx, t.cy, roi):
                    t.calc_speed(cam)

            # ── Crossing check: ALL active tracks, every frame ────────────────
            # Use self.trk.tracks (not the min_hits-filtered `tracks`) so that
            # vehicles crossing the line are detected even before they reach the
            # confirmed-track threshold.  The min_hits=2 guard inside
            # check_track() still rejects ghost boxes.
            # Events are returned so _poll() can accumulate from ALL frames,
            # even frames that get dropped during queue drain.
            crossings: list[dict] = []
            if cl and cl.active:
                for t in self.trk.tracks:
                    ev = cl.check_track(t, fid)
                    if ev:
                        crossings.append(ev)

            result = (enhanced, tracks, finished, crossings)
            try:
                self.out_q.put_nowait(result)
            except queue.Full:
                try:    self.out_q.get_nowait()
                except queue.Empty: pass
                try:    self.out_q.put_nowait(result)
                except queue.Full:  pass

    def stop(self):
        self._stop.set()


# ═══════════════════════════════════════════════════════════════════════════════
#  Drawing helpers
# ═══════════════════════════════════════════════════════════════════════════════
def speed_color(kmh: float, limit: float = 60) -> tuple:
    for thresh, col in SPEED_COLORS:
        if kmh < thresh:
            return col
    return SPEED_COLORS[-1][1]


def draw_tracks(frame: np.ndarray, tracks: list, scale: float,
                show_boxes: bool, show_speed: bool) -> np.ndarray:
    """Draw bounding boxes + ID badges only.
    Speed is NOT shown here — it is displayed via the flash overlay in _render()
    only at the moment of line crossing."""
    for t in tracks:
        x = int(t.bbox[0] * scale)
        y = int(t.bbox[1] * scale)
        w = int(t.bbox[2] * scale)
        h = int(t.bbox[3] * scale)
        clr = (0, 220, 80)   # neutral green for all non-crossed tracks

        # Bounding box
        if show_boxes:
            cv2.rectangle(frame, (x, y), (x+w, y+h), clr, 2, cv2.LINE_AA)

        if not show_speed:
            continue

        # ID badge (always shown when show_speed is True, so the user can read IDs)
        id_txt = f"#{t.id}"
        (iw, ih), _ = cv2.getTextSize(id_txt, cv2.FONT_HERSHEY_SIMPLEX, 0.50, 1)
        ix = x + w // 2 - iw // 2
        iy = max(ih + 6, y - 4)
        cv2.rectangle(frame, (ix-3, iy-ih-3), (ix+iw+3, iy+3), (0, 0, 0), -1)
        cv2.putText(frame, id_txt, (ix, iy),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.50, clr, 1, cv2.LINE_AA)

    return frame


def draw_cross_flash(frame: np.ndarray, flash: dict, scale: float,
                     speed_limit: float, now: float, ttl: float = 2.5):
    """Draw speed + direction overlay for recently-crossed vehicles.
    Each entry fades out after `ttl` seconds."""
    stale = []
    for tid, info in flash.items():
        age = now - info["t0"]
        if age > ttl:
            stale.append(tid)
            continue
        cx = int(info["cx"] * scale)
        cy = int(info["cy"] * scale)
        spd   = info["speed"]
        direc = info["dir"]
        clr   = (0, 255, 200) if direc == "IN" else (0, 80, 255)

        # Fade alpha: 1.0 → 0.0 over ttl seconds
        alpha = max(0.0, 1.0 - age / ttl)

        # Outer glow circle
        radius = int(30 + 10 * alpha)
        overlay = frame.copy()
        cv2.circle(overlay, (cx, cy), radius, clr, 2, cv2.LINE_AA)
        cv2.addWeighted(overlay, alpha * 0.6, frame, 1 - alpha * 0.6, 0, frame)

        # Direction + speed text
        spd_txt = f"{direc}  {spd:.0f} km/h" if spd > 0 else direc
        over_limit = spd > 0 and spd > speed_limit
        txt_clr = (0, 60, 255) if over_limit else clr
        (tw, th), _ = cv2.getTextSize(spd_txt, cv2.FONT_HERSHEY_SIMPLEX, 0.65, 2)
        tx = cx - tw // 2
        ty = cy - radius - 8
        cv2.rectangle(frame, (tx-5, ty-th-4), (tx+tw+5, ty+4), (0, 0, 0), -1)
        cv2.putText(frame, spd_txt, (tx, ty),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.65, txt_clr, 2, cv2.LINE_AA)
        if over_limit:
            cv2.putText(frame, "⚠ TEZKOR", (tx, ty + th + 14),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.45, (0, 40, 255), 1, cv2.LINE_AA)

    for tid in stale:
        del flash[tid]


# ═══════════════════════════════════════════════════════════════════════════════
#  Camera Calibrator (popup window)
# ═══════════════════════════════════════════════════════════════════════════════
class CalibratorWindow:
    """
    Popup kamera kalibrlash oynasi.
    Usullar:
      1. Avto (Hough vanishing point)
      2. Qo'lda 2+ parallel chiziq
      3. Ma'lum masofadagi 2 nuqta
    Saqlash tugmasi parent app ga params qaytaradi (callback).
    """
    def __init__(self, parent, callback=None, frame_ref=None):
        self.callback  = callback
        self.frame     = frame_ref.copy() if frame_ref is not None else None
        self.vw = self.vh = 0
        if self.frame is not None:
            self.vh, self.vw = self.frame.shape[:2]

        self.lines       = []
        self.drawing_line = False
        self.line_start  = None
        self._temp       = None
        self.ref_pts     = []
        self.setting_ref = False
        self.vp          = None
        self.pitch       = None
        self.focal       = None
        self.px_per_m    = None   # set by _calc_from_ref; passed to main app
        self.sx = self.sy = 1.0
        self.ox = self.oy = 0

        self.win = tk.Toplevel(parent)
        self.win.title("Kamera Kalibrlash")
        self.win.geometry("1050x650")
        self.win.configure(bg=BG_C)
        self._build()

    def _build(self):
        st = ttk.Style(); st.theme_use("clam")
        st.configure("TFrame",  background=BG_C)
        st.configure("TLabel",  background=BG_C, foreground="#b0b8c8",
                     font=("Consolas", 10))
        st.configure("TButton", font=("Consolas", 9, "bold"), padding=5)

        left = ttk.Frame(self.win)
        left.pack(side="left", fill="both", expand=True, padx=6, pady=6)
        self.canvas = tk.Canvas(left, bg="#000", highlightthickness=0, cursor="crosshair")
        self.canvas.pack(fill="both", expand=True)
        self.canvas.bind("<ButtonPress-1>", self._click)
        self.canvas.bind("<Motion>",        self._motion)
        self.canvas.bind("<Configure>",     lambda _: self._render())

        ctrl = ttk.Frame(left); ctrl.pack(fill="x", pady=3)
        ttk.Button(ctrl, text="📂 Rasm/Video kadr",
                   command=self._open).pack(side="left", padx=2)
        ttk.Button(ctrl, text="🔍 Avto-aniqlash",
                   command=self._auto_detect).pack(side="left", padx=2)
        ttk.Button(ctrl, text="📏 Qo'lda chiziq",
                   command=self._start_line).pack(side="left", padx=2)
        ttk.Button(ctrl, text="📐 2 nuqta+masofa",
                   command=self._start_ref).pack(side="left", padx=2)
        ttk.Button(ctrl, text="🗑 Tozalash",
                   command=self._clear).pack(side="left", padx=2)

        # Sidebar
        right = tk.Frame(self.win, bg=SB_C, width=300)
        right.pack(side="right", fill="y"); right.pack_propagate(False)
        tk.Label(right, text="KAMERA PARAMETRLARI", bg=SB_C, fg=ACC_C,
                 font=("Consolas", 11, "bold")).pack(pady=10)

        self.lbl_vp    = tk.Label(right, text="Vanishing Point: —",
                                  bg=SB_C, fg="#94a3b8", font=("Consolas", 9))
        self.lbl_pitch = tk.Label(right, text="Egilish burchagi: —",
                                  bg=SB_C, fg="#94a3b8", font=("Consolas", 9))
        self.lbl_focal = tk.Label(right, text="Fokus masofasi: —",
                                  bg=SB_C, fg="#94a3b8", font=("Consolas", 9))
        for lbl in (self.lbl_vp, self.lbl_pitch, self.lbl_focal):
            lbl.pack(anchor="w", padx=14, pady=2)

        tk.Frame(right, bg="#1e293b", height=1).pack(fill="x", padx=12, pady=8)
        tk.Label(right, text="QO'LDA KIRITISH", bg=SB_C, fg="#475569",
                 font=("Consolas", 8, "bold")).pack(anchor="w", padx=12)

        fields = [
            ("Kamera balandligi (m):", "e_cam_h", "6.0"),
            ("Fokus taxmini (px):",    "e_cam_f", "800"),
            ("Sensor kengligi (mm):",  "e_sensor_w", "6.17"),
        ]
        for lbl_text, attr, default in fields:
            fr = tk.Frame(right, bg=SB_C); fr.pack(fill="x", padx=14, pady=2)
            tk.Label(fr, text=lbl_text, bg=SB_C, fg="#94a3b8",
                     font=("Consolas", 9)).pack(side="left")
            e = ttk.Entry(fr, width=8); e.pack(side="right")
            e.insert(0, default)
            setattr(self, attr, e)

        ttk.Button(right, text="Burchakni hisoblash",
                   command=self._calc_pitch).pack(fill="x", padx=14, pady=6)

        tk.Frame(right, bg="#1e293b", height=1).pack(fill="x", padx=12, pady=6)
        tk.Label(right, text="REFERENS MASOFA", bg=SB_C, fg="#475569",
                 font=("Consolas", 8, "bold")).pack(anchor="w", padx=12)
        fr2 = tk.Frame(right, bg=SB_C); fr2.pack(fill="x", padx=14, pady=2)
        tk.Label(fr2, text="Haqiqiy masofa (m):", bg=SB_C, fg="#94a3b8",
                 font=("Consolas", 9)).pack(side="left")
        self.e_ref_dist = ttk.Entry(fr2, width=8); self.e_ref_dist.pack(side="right")
        self.e_ref_dist.insert(0, "5.0")

        tk.Frame(right, bg="#1e293b", height=1).pack(fill="x", padx=12, pady=6)
        ttk.Button(right, text="💾 JSON saqlash",
                   command=self._save_json).pack(fill="x", padx=14, pady=2)
        ttk.Button(right, text="✅ Asosiy dasturga qo'llash",
                   command=self._apply_to_main).pack(fill="x", padx=14, pady=4)

        self.lbl_result = tk.Label(right, text="", bg=SB_C, fg="#34d399",
                                   font=("Consolas", 9), wraplength=260, justify="left")
        self.lbl_result.pack(anchor="w", padx=14, pady=6)

        if self.frame is not None:
            self._render()

    # ── I/O ──────────────────────────────────────────────────────────────────
    def _open(self):
        p = filedialog.askopenfilename(
            filetypes=[("Rasm/Video", "*.jpg *.jpeg *.png *.bmp *.mp4 *.avi *.mov"),
                       ("All", "*.*")],
            parent=self.win)
        if not p:
            return
        if p.lower().endswith(('.mp4', '.avi', '.mov', '.mkv')):
            cap = cv2.VideoCapture(p)
            cap.set(cv2.CAP_PROP_POS_FRAMES, int(cap.get(cv2.CAP_PROP_FRAME_COUNT)) // 2)
            ok, frame = cap.read(); cap.release()
            if ok: self.frame = frame
        else:
            self.frame = cv2.imread(p)
        if self.frame is not None:
            self.vh, self.vw = self.frame.shape[:2]
            self.lines = []; self.ref_pts = []; self.vp = None
            self._render()

    def _render(self):
        if self.frame is None: return
        disp = self.frame.copy()
        cw = max(self.canvas.winfo_width(), 100)
        ch = max(self.canvas.winfo_height(), 100)
        r  = min(cw / max(self.vw, 1), ch / max(self.vh, 1))
        nw, nh = int(self.vw * r), int(self.vh * r)
        self.sx, self.sy = self.vw / max(nw, 1), self.vh / max(nh, 1)
        self.ox, self.oy = (cw - nw) // 2, (ch - nh) // 2

        for i, (x1, y1, x2, y2) in enumerate(self.lines):
            cv2.line(disp, (x1,y1), (x2,y2), (0,255,255), 2, cv2.LINE_AA)
            cv2.putText(disp, str(i+1), (x1+5,y1-5),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0,255,255), 1)

        if self.drawing_line and self.line_start and self._temp:
            cv2.line(disp, tuple(self.line_start), tuple(self._temp), (0,200,200), 1)

        for i, pt in enumerate(self.ref_pts):
            cv2.circle(disp, tuple(pt), 8, (255,100,100), -1)
            cv2.putText(disp, f"R{i+1}", (pt[0]+10, pt[1]),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.5, (255,100,100), 1)
        if len(self.ref_pts) == 2:
            cv2.line(disp, tuple(self.ref_pts[0]), tuple(self.ref_pts[1]),
                     (255,100,100), 2)

        if self.vp:
            vx, vy = int(self.vp[0]), int(self.vp[1])
            if 0 <= vx < self.vw and 0 <= vy < self.vh:
                cv2.circle(disp, (vx, vy), 12, (0,0,255), 3)
                cv2.putText(disp, "VP", (vx+15, vy),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0,0,255), 2)
            cv2.line(disp, (0, vy), (self.vw, vy), (0,0,200), 1, cv2.LINE_AA)

        rgb = cv2.cvtColor(disp, cv2.COLOR_BGR2RGB)
        self._tk = ImageTk.PhotoImage(Image.fromarray(cv2.resize(rgb, (nw, nh))))
        self.canvas.delete("all")
        self.canvas.create_image(self.ox, self.oy, anchor="nw", image=self._tk)

    def _c2v(self, ex, ey):
        return (max(0, min(self.vw-1, int((ex-self.ox)*self.sx))),
                max(0, min(self.vh-1, int((ey-self.oy)*self.sy))))

    def _click(self, e):
        if self.frame is None: return
        vx, vy = self._c2v(e.x, e.y)
        if self.drawing_line:
            if self.line_start is None:
                self.line_start = [vx, vy]
            else:
                self.lines.append((self.line_start[0], self.line_start[1], vx, vy))
                self.line_start = None; self.drawing_line = False
                self._render()
                if len(self.lines) >= 2:
                    self._calc_vp()
        elif self.setting_ref:
            self.ref_pts.append([vx, vy])
            if len(self.ref_pts) >= 2:
                self.setting_ref = False
                self._calc_from_ref()
            self._render()

    def _motion(self, e):
        if self.drawing_line and self.line_start and self.frame is not None:
            self._temp = list(self._c2v(e.x, e.y))
            self._render()

    def _start_line(self):
        self.drawing_line = True; self.line_start = None

    def _start_ref(self):
        self.setting_ref = True; self.ref_pts = []

    def _clear(self):
        self.lines = []; self.ref_pts = []; self.vp = None
        self.pitch = self.focal = None
        self.lbl_vp.configure(text="VP: —")
        self.lbl_pitch.configure(text="Egilish burchagi: —")
        self.lbl_focal.configure(text="Fokus masofasi: —")
        self.lbl_result.configure(text="")
        if self.frame is not None: self._render()

    # ── Detection ─────────────────────────────────────────────────────────────
    def _auto_detect(self):
        if self.frame is None: return
        gray  = cv2.cvtColor(self.frame, cv2.COLOR_BGR2GRAY)
        blur  = cv2.GaussianBlur(gray, (5,5), 0)
        edges = cv2.Canny(blur, 50, 150)
        mask  = np.zeros_like(edges)
        mask[self.vh//3:, :] = 255
        edges = cv2.bitwise_and(edges, mask)
        lines = cv2.HoughLinesP(edges, 1, np.pi/180, threshold=80,
                                minLineLength=80, maxLineGap=30)
        if lines is None:
            messagebox.showinfo("Info", "Chiziqlar topilmadi. Qo'lda chizing.",
                                parent=self.win); return
        good = []
        for l in lines:
            x1,y1,x2,y2 = l[0]
            angle = abs(np.degrees(np.arctan2(y2-y1, x2-x1)))
            if 15 < angle < 85:
                good.append((x1,y1,x2,y2))
        if len(good) < 2:
            messagebox.showinfo("Info", f"Faqat {len(good)} chiziq topildi.",
                                parent=self.win); return
        good.sort(key=lambda l: (l[2]-l[0])**2+(l[3]-l[1])**2, reverse=True)
        self.lines = good[:6]
        self._render(); self._calc_vp()

    def _calc_vp(self):
        if len(self.lines) < 2: return
        pts = []
        for i in range(len(self.lines)):
            for j in range(i+1, len(self.lines)):
                pt = self._intersect(self.lines[i], self.lines[j])
                if pt and (-self.vw < pt[0] < 2*self.vw) and (-self.vh < pt[1] < 2*self.vh):
                    pts.append(pt)
        if not pts:
            messagebox.showinfo("Info", "Kesishish topilmadi.", parent=self.win); return
        self.vp = (np.median([p[0] for p in pts]), np.median([p[1] for p in pts]))
        self._calc_pitch(); self._render()

    @staticmethod
    def _intersect(l1, l2):
        x1,y1,x2,y2 = l1; x3,y3,x4,y4 = l2
        d = (x1-x2)*(y3-y4) - (y1-y2)*(x3-x4)
        if abs(d) < 1e-6: return None
        t = ((x1-x3)*(y3-y4) - (y1-y3)*(x3-x4)) / d
        return (x1 + t*(x2-x1), y1 + t*(y2-y1))

    def _calc_pitch(self):
        if self.vp is None: return
        try:
            f = float(self.e_cam_f.get())
        except ValueError:
            f = 800.0
        cy           = self.vh / 2.0
        self.pitch   = np.degrees(np.arctan((cy - self.vp[1]) / f))
        self.focal   = f
        self.lbl_vp.configure(text=f"VP: ({self.vp[0]:.0f}, {self.vp[1]:.0f})")
        self.lbl_pitch.configure(text=f"Egilish: {self.pitch:.1f}°", fg="#34d399")
        self.lbl_focal.configure(text=f"Fokus: {f:.0f} px")
        self.lbl_result.configure(
            text=f"Egilish: {self.pitch:.1f}°\nVP: ({self.vp[0]:.0f},{self.vp[1]:.0f})\n"
                 f"Fokus: {f:.0f} px")

    def _calc_from_ref(self):
        if len(self.ref_pts) < 2: return
        try:
            real_d = float(self.e_ref_dist.get())
            H      = float(self.e_cam_h.get())
            f      = float(self.e_cam_f.get())
        except ValueError:
            messagebox.showwarning("Xato", "Qiymatlarni to'g'ri kiriting!",
                                   parent=self.win); return
        p1, p2 = self.ref_pts
        px_d   = np.sqrt((p2[0]-p1[0])**2 + (p2[1]-p1[1])**2)
        cy     = self.vh / 2.0
        avg_y  = (p1[1]+p2[1]) / 2.0
        self.pitch   = np.degrees(np.arctan(H/real_d) + np.arctan((cy-avg_y)/f))
        self.focal   = f
        self.px_per_m = px_d / real_d   # measured scale, fed back to camera model
        self.lbl_pitch.configure(text=f"Egilish: {self.pitch:.1f}° (taxminiy)", fg="#fbbf24")
        self.lbl_result.configure(
            text=f"Piksel/metr: {self.px_per_m:.1f}\nEgilish: {self.pitch:.1f}°\n"
                 f"H={H}m f={f}px\nBu px/m qo'llandi ✓")
        self._render()

    def _save_json(self):
        try:   H = float(self.e_cam_h.get())
        except: H = 6.0
        try:   f = float(self.e_cam_f.get())
        except: f = 800.0
        data = {
            "camera_height_m":  H,
            "camera_pitch_deg": round(self.pitch, 2) if self.pitch else 0,
            "focal_length_px":  round(f, 1),
            "frame_width":  self.vw,
            "frame_height": self.vh,
            "vanishing_point": [round(self.vp[0],1), round(self.vp[1],1)] if self.vp else None,
        }
        path = filedialog.asksaveasfilename(
            defaultextension=".json", filetypes=[("JSON","*.json")],
            initialfile="camera_params.json", parent=self.win)
        if path:
            with open(path, "w") as fp:
                json.dump(data, fp, indent=2)
            messagebox.showinfo("Saqlandi!", path, parent=self.win)

    def _apply_to_main(self):
        if self.pitch is None:
            messagebox.showwarning("Hisoblang", "Avval burchakni hisoblang!",
                                   parent=self.win); return
        if self.callback:
            try:   H = float(self.e_cam_h.get())
            except: H = 6.0
            try:   f = float(self.e_cam_f.get())
            except: f = 800.0
            params = {
                "camera_height_m":  H,
                "camera_pitch_deg": self.pitch,
                "focal_length_px":  f,
                "frame_width":  self.vw  if self.vw  else 1280,
                "frame_height": self.vh  if self.vh  else  720,
            }
            if self.px_per_m is not None:
                params["px_per_m"] = self.px_per_m
            self.callback(params)
        messagebox.showinfo("Qo'llandi", "Kamera parametrlari asosiy dasturga qo'llandi!",
                            parent=self.win)
        self.win.destroy()


# ═══════════════════════════════════════════════════════════════════════════════
#  CrossLine — virtual counting line
# ═══════════════════════════════════════════════════════════════════════════════
class CrossLine:
    """
    Virtual counting line. Vehicles crossing it are counted once per direction.

    Side convention:
      sign = (x2-x1)*(py-y1) - (y2-y1)*(px-x1)
      sign > 0  → side A  (counted as "IN"  when crossing A→B)
      sign < 0  → side B  (counted as "OUT" when crossing B→A)
    """

    def __init__(self):
        self.p1: list | None = None   # [x, y]
        self.p2: list | None = None   # [x, y]
        self.active = False
        self.count_in  = 0
        self.count_out = 0
        self.events: list[dict] = []  # [{id, dir, frame, speed}]

    def define(self, p1: list, p2: list):
        self.p1 = p1[:]
        self.p2 = p2[:]
        self.active = True

    def reset_counts(self):
        self.count_in = self.count_out = 0
        self.events.clear()

    def clear(self):
        self.p1 = self.p2 = None
        self.active = False
        self.reset_counts()

    @staticmethod
    def _side(px, py, x1, y1, x2, y2) -> float:
        return (x2 - x1) * (py - y1) - (y2 - y1) * (px - x1)

    def check_track(self, track: "Track", fid: int) -> dict | None:
        """
        Call inside the Worker on EVERY frame for EVERY active track.
        Returns crossing event dict {id, dir, frame, speed} or None.
        Counts NOT modified here — caller (_poll) accumulates them so that
        events from skipped (queue-drained) frames are never lost.
        Each track is counted only ONCE via track.crossed flag.

        Direction convention (matches the drawn arrow):
          Arrow points to the LEFT of the line (p1→p2) = positive-side = side A.
          Crossing TO side A (curr_side > 0)  → "IN"   (moving with arrow)
          Crossing TO side B (curr_side < 0)  → "OUT"  (moving against arrow)
        """
        # min_hits=2: accept after 2 consistent detections (avoids ghost boxes)
        if not self.active or track.crossed or track.hits < 2:
            return None
        x1, y1 = self.p1
        x2, y2 = self.p2

        curr_side = self._side(track.cx,      track.cy,      x1, y1, x2, y2)
        prev_side = self._side(track.prev_cx,  track.prev_cy, x1, y1, x2, y2)

        # Initialise side reference on first appearance — no crossing yet
        if track.line_side is None:
            track.line_side = curr_side
            return None

        # Require a real movement step to suppress jitter-triggered false crossings
        motion = np.hypot(track.cx - track.prev_cx, track.cy - track.prev_cy)
        if motion < 2.0:
            track.line_side = curr_side
            return None

        # True crossing: centroid moved from one side to the other
        if prev_side * curr_side < 0:
            direction       = "IN" if curr_side > 0 else "OUT"
            track.crossed   = True
            track.cross_dir = direction
            track.line_side = curr_side
            return {"id":    track.id,  "dir":   direction,
                    "frame": fid,       "speed": round(track.speed_kmh, 1),
                    "cx":    track.cx,  "cy":    track.cy}

        track.line_side = curr_side
        return None

    def draw(self, frame: np.ndarray, scale: float = 1.0):
        """Draw the counting line + counts onto frame."""
        if not self.active or self.p1 is None or self.p2 is None:
            return
        x1, y1 = int(self.p1[0] * scale), int(self.p1[1] * scale)
        x2, y2 = int(self.p2[0] * scale), int(self.p2[1] * scale)

        # Gradient line: yellow core
        cv2.line(frame, (x1, y1), (x2, y2), (0, 0, 0),       5, cv2.LINE_AA)
        cv2.line(frame, (x1, y1), (x2, y2), (0, 255, 255),    2, cv2.LINE_AA)

        # Endpoint dots
        cv2.circle(frame, (x1, y1), 5, (0, 200, 255), -1)
        cv2.circle(frame, (x2, y2), 5, (0, 200, 255), -1)

        # Arrow showing "IN" direction (left normal of line)
        mx, my = (x1+x2)//2, (y1+y2)//2
        nx =  (y2 - y1); ny = -(x2 - x1)
        mag = max(np.hypot(nx, ny), 1e-6)
        arlen = 22
        ax, ay = int(mx + arlen * nx / mag), int(my + arlen * ny / mag)
        cv2.arrowedLine(frame, (mx, my), (ax, ay), (0, 255, 255), 2,
                        cv2.LINE_AA, tipLength=0.4)

        # Count badge
        txt = f"IN:{self.count_in}  OUT:{self.count_out}"
        (tw, th), _ = cv2.getTextSize(txt, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 2)
        bx, by = mx - tw // 2, my - 18
        cv2.rectangle(frame, (bx-5, by-th-4), (bx+tw+5, by+4), (0, 0, 0), -1)
        cv2.rectangle(frame, (bx-5, by-th-4), (bx+tw+5, by+4), (0, 200, 255), 1)
        cv2.putText(frame, txt, (bx, by),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 255), 2, cv2.LINE_AA)


# ═══════════════════════════════════════════════════════════════════════════════
#  Main Application
# ═══════════════════════════════════════════════════════════════════════════════
class SpeedAnalyzerPro:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Speed Analyzer Pro")
        self.root.geometry("1400x820")
        self.root.configure(bg=BG_C)

        # Video state
        self.cap        = None
        self.fps        = 30.0
        self.video_name = ""
        self.playing    = False
        self.worker     = None
        self.cap_thread = None

        self.in_q:  queue.Queue = queue.Queue(maxsize=CAP_Q_SIZE)
        self.out_q: queue.Queue = queue.Queue(maxsize=RES_Q_SIZE)

        self._last_frame : np.ndarray | None = None
        self._last_tracks: list              = []
        self._ftimes = deque(maxlen=30)

        # ROI
        self.roi      = []
        self._drawing_roi = False
        self._roi_tmp     = None
        self._roi_mask: np.ndarray | None = None

        # Scale/offset for canvas→video mapping
        self.sx = self.sy = 1.0
        self.ox = self.oy = 0

        # Objects
        self.model, self.model_name = load_yolo()
        self.yolo_det  = YOLODetector(self.model) if self.model else None
        self.bg_det    = BGDetector()
        self.enhancer  = Enhancer()
        self.stabilizer = VideoStabilizer()
        self.cam       = CameraModel()
        self.tracker   = Tracker(30.0, self.cam)

        # Results
        self.results      = []
        self.recorded_ids = set()
        self.ssdir        = "speed_screenshots"
        os.makedirs(self.ssdir, exist_ok=True)
        self._slider_busy = False   # recursion guard for slider.set ↔ _seek

        # Counting line
        self.cross_line    = CrossLine()
        self._drawing_line = False
        self._line_tmp: list | None = None
        # Flash overlays: {track_id: {speed, dir, cx, cy, t0}}  (shown for 2.5 s)
        self._cross_flash: dict = {}

        # UI vars
        self.var_limit   = tk.DoubleVar(value=60.0)
        self.var_stab    = tk.BooleanVar(value=True)
        self.var_stabview = tk.BooleanVar(value=True)
        self.var_cam_on  = tk.BooleanVar(value=False)
        self.var_cam_H   = tk.DoubleVar(value=6.0)
        self.var_cam_T   = tk.DoubleVar(value=15.0)
        self.var_cam_P   = tk.DoubleVar(value=0.0)
        self.var_cam_F   = tk.DoubleVar(value=800.0)
        self.var_min_a   = tk.IntVar(value=1500)
        self.var_max_a   = tk.IntVar(value=150000)
        self.var_trail   = tk.BooleanVar(value=True)
        self.var_boxes   = tk.BooleanVar(value=True)
        self.var_speed_lbl = tk.BooleanVar(value=True)
        self.var_autosave  = tk.BooleanVar(value=True)
        self.var_speed_mul = tk.DoubleVar(value=1.0)
        self.var_speed_k   = tk.DoubleVar(value=1.0)   # speed calibration factor

        self._build_ui()

    # ─── UI ─────────────────────────────────────────────────────────────────
    def _build_ui(self):
        st = ttk.Style(); st.theme_use("clam")
        st.configure("TFrame",       background=BG_C)
        st.configure("TLabel",       background=BG_C, foreground="#94a3b8",
                     font=("Consolas", 10))
        st.configure("TButton",      font=("Consolas", 9, "bold"), padding=5)
        st.configure("TCheckbutton", background=SB_C, foreground="#b0b8c8",
                     font=("Consolas", 9))

        # ── Header ──────────────────────────────────────────────────────────
        hdr = tk.Frame(self.root, bg=SB_C, height=44)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Label(hdr, text=f"Speed Analyzer Pro  [{self.model_name or 'BG mode'}]",
                 bg=SB_C, fg=ACC_C, font=("Consolas",13,"bold")).pack(side="left", padx=12)
        self.lbl_fps   = tk.Label(hdr, text="", bg=SB_C, fg="#475569",
                                  font=("Consolas",10))
        self.lbl_fps.pack(side="right", padx=12)
        self.lbl_info  = tk.Label(hdr, text="", bg=SB_C, fg="#34d399",
                                  font=("Consolas",11,"bold"))
        self.lbl_info.pack(side="right", padx=14)
        self.lbl_stab  = tk.Label(hdr, text="", bg=SB_C, fg="#475569",
                                  font=("Consolas",9))
        self.lbl_stab.pack(side="right", padx=12)

        main = ttk.Frame(self.root)
        main.pack(fill="both", expand=True)

        # ── Left: canvas + controls ─────────────────────────────────────────
        left = ttk.Frame(main)
        left.pack(side="left", fill="both", expand=True, padx=4, pady=4)

        self.canvas = tk.Canvas(left, bg="#000", highlightthickness=0, cursor="crosshair")
        self.canvas.pack(fill="both", expand=True)
        self.canvas.bind("<Configure>",
                         lambda _: self._render() if self._last_frame is not None else None)
        self.canvas.bind("<ButtonPress-1>", self._canvas_click)
        self.canvas.bind("<Motion>",        self._canvas_motion)
        self.canvas.bind("<ButtonPress-3>", self._canvas_rclick)

        # Row 1: file + playback
        row1 = tk.Frame(left, bg=BG_C); row1.pack(fill="x", pady=2)
        ttk.Button(row1, text="📂 Video", command=self._open_video).pack(side="left", padx=2)
        self.btn_play = ttk.Button(row1, text="▶ Play", command=self._toggle_play)
        self.btn_play.pack(side="left", padx=2)
        for txt, n in [("◀◀",-5),("◀",-1),("▶",1),("▶▶",5)]:
            ttk.Button(row1, text=txt, width=3,
                       command=lambda n=n: self._step(n)).pack(side="left", padx=1)
        self.slider = ttk.Scale(row1, from_=0, to=100, orient="horizontal",
                                command=self._seek)
        self.slider.pack(side="left", fill="x", expand=True, padx=4)
        self.lbl_time = ttk.Label(row1, text="0.0s")
        self.lbl_time.pack(side="right", padx=4)

        # Row 2: tools
        row2 = tk.Frame(left, bg=BG_C); row2.pack(fill="x", pady=2)
        ttk.Button(row2, text="🔷 ROI belgilash",
                   command=self._roi_start).pack(side="left", padx=2)
        ttk.Button(row2, text="📷 Kamera kalibrlash",
                   command=self._open_calibrator).pack(side="left", padx=2)
        ttk.Button(row2, text="📂 JSON import",
                   command=self._load_cam_json).pack(side="left", padx=2)
        ttk.Button(row2, text="〰 Hisoblash chizig'i",
                   command=self._line_start).pack(side="left", padx=2)
        ttk.Button(row2, text="🗑 Tozalash",
                   command=self._clear_all).pack(side="left", padx=2)

        # ── Right: scrollable sidebar ────────────────────────────────────────
        sb_outer = tk.Frame(main, bg=SB_C, width=240)
        sb_outer.pack(side="right", fill="y"); sb_outer.pack_propagate(False)
        sb_cv = tk.Canvas(sb_outer, bg=SB_C, highlightthickness=0)
        sb_sc = ttk.Scrollbar(sb_outer, orient="vertical", command=sb_cv.yview)
        self.sb = tk.Frame(sb_cv, bg=SB_C)
        self.sb.bind("<Configure>",
                     lambda _: sb_cv.configure(scrollregion=sb_cv.bbox("all")))
        sb_cv.create_window((0,0), window=self.sb, anchor="nw")
        sb_cv.configure(yscrollcommand=sb_sc.set)
        sb_cv.pack(side="left", fill="both", expand=True)
        sb_sc.pack(side="right", fill="y")

        def sec(t):
            tk.Label(self.sb, text=t, bg=SB_C, fg="#475569",
                     font=("Consolas",8,"bold")).pack(anchor="w", padx=10, pady=(10,2))
        def hr():
            tk.Frame(self.sb, bg="#1e293b", height=1).pack(fill="x", padx=10, pady=4)
        def chk(text, var, cmd=None):
            ttk.Checkbutton(self.sb, text=text, variable=var,
                            command=cmd).pack(anchor="w", padx=12)
        def param(label, var):
            fr = tk.Frame(self.sb, bg=SB_C); fr.pack(fill="x", padx=12, pady=1)
            tk.Label(fr, text=label, bg=SB_C, fg="#94a3b8",
                     font=("Consolas",9)).pack(side="left")
            ttk.Entry(fr, textvariable=var, width=8).pack(side="right")
        def btn(text, cmd):
            ttk.Button(self.sb, text=text, command=cmd).pack(fill="x", padx=12, pady=2)

        # STATUS
        sec("HOLAT")
        self.ls_roi   = tk.Label(self.sb, text="ROI: ✗", bg=SB_C, fg="#f87171",
                                 font=("Consolas",9))
        self.ls_roi.pack(anchor="w", padx=14)
        self.ls_cam   = tk.Label(self.sb, text="Kamera: —", bg=SB_C, fg="#fbbf24",
                                 font=("Consolas",9))
        self.ls_cam.pack(anchor="w", padx=14)
        self.ls_trk   = tk.Label(self.sb, text="Treklar: 0", bg=SB_C, fg="#94a3b8",
                                 font=("Consolas",9))
        self.ls_trk.pack(anchor="w", padx=14)
        self.ls_stab2 = tk.Label(self.sb, text="Stab: —", bg=SB_C, fg="#94a3b8",
                                 font=("Consolas",9))
        self.ls_stab2.pack(anchor="w", padx=14)
        hr()

        # MODEL
        sec("MODEL")
        btn("Boshqa model yuklash", self._load_model)
        self.lbl_model = tk.Label(
            self.sb, text=self.model_name or "BG mode",
            bg=SB_C, fg="#34d399" if self.model else "#f87171",
            font=("Consolas",8), wraplength=200)
        self.lbl_model.pack(anchor="w", padx=12)
        hr()

        # STABILIZATSIYA
        sec("VIDEO STABILIZATSIYA")
        chk("Yoqish",          self.var_stab,     self._toggle_stab)
        chk("Kadrni rostlash", self.var_stabview,  self._toggle_stab)
        hr()

        # TEZLIK CHEGARASI
        sec("TEZLIK CHEGARASI")
        param("Chegara (km/h):", self.var_limit)
        hr()

        # KAMERA
        sec("KAMERA MODELI")
        chk("Perspektiv tuzatish", self.var_cam_on, self._apply_cam)
        param("Balandlik (m):",    self.var_cam_H)
        param("Tilt (°):",         self.var_cam_T)
        param("Pan (°):",          self.var_cam_P)
        param("Fokus (px):",       self.var_cam_F)
        param("Tezlik koeff.:",    self.var_speed_k)
        btn("Qo'llash", self._apply_cam)
        hr()

        # DETEKSIYA
        sec("DETEKSIYA")
        param("Min area (px²):", self.var_min_a)
        param("Max area (px²):", self.var_max_a)
        hr()

        # KO'RINISH
        sec("KO'RINISH")
        chk("Bounding box", self.var_boxes)
        chk("Yo'l izi",      self.var_trail)
        chk("Tezlik yorliq", self.var_speed_lbl)
        fr = tk.Frame(self.sb, bg=SB_C); fr.pack(fill="x", padx=12, pady=2)
        tk.Label(fr, text="Tezlik:", bg=SB_C, fg="#94a3b8",
                 font=("Consolas",9)).pack(side="left")
        for s, lbl in [(.5,"½"),(1,"1x"),(2,"2x"),(4,"4x")]:
            ttk.Button(fr, text=lbl, width=3,
                       command=lambda s=s: self.var_speed_mul.set(s)).pack(side="left", padx=1)
        hr()

        # EKSPORT
        sec("EKSPORT")
        btn("📥 Excel eksport", self._export_excel)
        chk("Avtomatik saqlash", self.var_autosave)
        btn("📸 Screenshot",     self._screenshot)
        hr()

        # STATISTIKA
        sec("STATISTIKA")
        self.lbl_total = tk.Label(self.sb, text="Jami: 0 avto", bg=SB_C, fg="#34d399",
                                  font=("Consolas",10,"bold"))
        self.lbl_total.pack(anchor="w", padx=14)
        self.lbl_avg   = tk.Label(self.sb, text="O'rtacha: —", bg=SB_C, fg="#94a3b8",
                                  font=("Consolas",9))
        self.lbl_avg.pack(anchor="w", padx=14)
        self.lbl_max_s = tk.Label(self.sb, text="Maks: —", bg=SB_C, fg="#94a3b8",
                                  font=("Consolas",9))
        self.lbl_max_s.pack(anchor="w", padx=14)
        self.lbl_over  = tk.Label(self.sb, text="Tezkor: 0", bg=SB_C, fg="#f87171",
                                  font=("Consolas",9))
        self.lbl_over.pack(anchor="w", padx=14)
        self.lbl_cross = tk.Label(self.sb, text="Chiziq: IN 0 | OUT 0",
                                  bg=SB_C, fg="#00e5ff", font=("Consolas",9,"bold"))
        self.lbl_cross.pack(anchor="w", padx=14)
        hr()

        # NATIJALAR
        sec("NATIJALAR")
        self.rl = tk.Listbox(self.sb, bg="#0f1318", fg="#c8cdd8",
                             font=("Consolas",8), selectbackground="#2563eb",
                             highlightthickness=0, bd=0, height=10)
        self.rl.pack(fill="x", padx=12, pady=4)

    # ─── Sidebar callbacks ──────────────────────────────────────────────────
    def _toggle_stab(self):
        self.stabilizer.enabled        = self.var_stab.get()
        self.stabilizer.stabilize_view = self.var_stabview.get()

    def _apply_cam(self):
        self.cam.speed_factor = max(0.01, self.var_speed_k.get())
        if self.var_cam_on.get():
            fw = int(self.cap.get(cv2.CAP_PROP_FRAME_WIDTH))  if self.cap else 1280
            fh = int(self.cap.get(cv2.CAP_PROP_FRAME_HEIGHT)) if self.cap else  720
            self.cam.setup(self.var_cam_H.get(), self.var_cam_T.get(),
                           self.var_cam_P.get(), self.var_cam_F.get(), fw, fh)
        else:
            self.cam.enabled = False
        self._update_status()

    def _load_model(self):
        path = filedialog.askopenfilename(
            filetypes=[("Model","*.pt *.onnx *.engine"),("All","*.*")])
        if not path: return
        self._stop_play()
        m, name = load_yolo(path)
        if m:
            self.model = m; self.model_name = name
            self.yolo_det = YOLODetector(m)
            self.lbl_model.configure(text=name, fg="#34d399")
        else:
            self.lbl_model.configure(text="Yuklab bo'lmadi", fg="#f87171")

    def _load_cam_json(self):
        path = filedialog.askopenfilename(filetypes=[("JSON","*.json")])
        if not path: return
        try:
            self.cam.load_json(path)
            self.var_cam_H.set(self.cam.H)
            self.var_cam_T.set(self.cam.tilt)
            self.var_cam_P.set(self.cam.pan)
            self.var_cam_F.set(self.cam.f)
            self.var_cam_on.set(True)
            self._update_status()
        except Exception as e:
            messagebox.showerror("Xato", str(e))

    def _open_calibrator(self):
        frame_ref = self._last_frame if self._last_frame is not None else None
        CalibratorWindow(self.root, callback=self._on_calib_result,
                         frame_ref=frame_ref)

    def _on_calib_result(self, params: dict):
        self.cam.H    = params.get("camera_height_m",  self.cam.H)
        self.cam.tilt = params.get("camera_pitch_deg", self.cam.tilt)
        self.cam.f    = params.get("focal_length_px",  self.cam.f)
        fw = params.get("frame_width",  1280)
        fh = params.get("frame_height",  720)
        self.cam.cx, self.cam.cy = fw/2.0, fh/2.0
        if "px_per_m" in params:
            self.cam.px_per_m = params["px_per_m"]
        self.cam.enabled = True
        self.var_cam_on.set(True)
        self.var_cam_H.set(self.cam.H)
        self.var_cam_T.set(self.cam.tilt)
        self.var_cam_F.set(self.cam.f)
        self._update_status()

    # ─── Video open ─────────────────────────────────────────────────────────
    def _open_video(self):
        path = filedialog.askopenfilename(
            filetypes=[("Video","*.mp4 *.avi *.mov *.mkv"),("All","*.*")])
        if not path: return
        self._stop_play()
        if self.cap: self.cap.release()
        self.cap = cv2.VideoCapture(path)
        self.fps = self.cap.get(cv2.CAP_PROP_FPS) or 30.0
        self.video_name = os.path.basename(path)
        total = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))
        self.slider.configure(to=max(1, total-1))
        self.tracker = Tracker(self.fps, self.cam)
        self.results = []; self.recorded_ids = set(); self._cross_flash = {}
        self.roi = []; self._roi_mask = None; self.rl.delete(0, "end")
        ok, f = self.cap.read()
        if ok:
            self._last_frame  = f
            self._last_tracks = []
            self._render()
        self.root.title(f"Speed Analyzer Pro — {self.video_name}")
        self._update_status()

    # ─── Playback ────────────────────────────────────────────────────────────
    def _toggle_play(self):
        if not self.cap: return
        if self.playing:
            self._stop_play()
        else:
            self._start_play()

    def _start_play(self):
        self.playing = True
        self.btn_play.configure(text="⏸ Pauza")
        for q in (self.in_q, self.out_q):
            while not q.empty():
                try: q.get_nowait()
                except queue.Empty: break
        if self.tracker is None:
            self.tracker = Tracker(self.fps, self.cam)
        self.tracker.reset()
        self.stabilizer.reset()
        self.bg_det = BGDetector(self.var_min_a.get(), self.var_max_a.get())
        self.worker = Worker(self.yolo_det, self.bg_det, self.enhancer,
                             self.tracker, self.in_q, self.out_q,
                             cross_line=self.cross_line)
        if self._roi_mask is not None:
            self.worker.set_roi(self._roi_mask)
        self.worker.start()
        self.cap_thread = threading.Thread(target=self._cap_loop, daemon=True)
        self.cap_thread.start()
        self._schedule_poll()

    def _stop_play(self):
        self.playing = False
        self.btn_play.configure(text="▶ Play")
        try: self.in_q.put_nowait(None)
        except queue.Full: pass
        if self.worker:
            self.worker.stop(); self.worker = None
        # Auto-save when stopping
        if self.var_autosave.get() and self.results:
            self._do_save_excel(auto=True)

    def _cap_loop(self):
        interval = 1.0 / self.fps
        t_next   = time.perf_counter() + interval
        while self.playing:
            ok, frame = self.cap.read()
            if not ok:
                try: self.in_q.put(None, timeout=1.0)
                except queue.Full: pass
                break
            # Apply stabilization in the capture thread
            stab_frame, dx, dy, _conf = self.stabilizer.process(frame)
            self._stab_dx, self._stab_dy = dx, dy
            try:
                self.in_q.put_nowait((int(self.cap.get(cv2.CAP_PROP_POS_FRAMES)), stab_frame))
            except queue.Full:
                pass
            sleep = t_next - time.perf_counter()
            if sleep > 0:
                time.sleep(sleep)
            t_next += interval

    def _schedule_poll(self):
        delay = max(8, int(1000.0 / self.fps))
        self.root.after(delay, self._poll)

    def _poll(self):
        if not self.playing:
            return

        # Drain the entire queue.
        # We render only the LATEST frame, but we MUST process crossing events
        # from EVERY frame — otherwise intermediate crossings are silently lost.
        latest_frame   = None
        latest_tracks  = None
        all_finished:  list = []
        all_crossings: list = []   # accumulated from ALL frames in queue

        try:
            while True:
                item = self.out_q.get_nowait()
                if item is None:
                    self._stop_play()
                    return
                frame_d, tracks_d, finished_d, crossings_d = item
                latest_frame  = frame_d
                latest_tracks = tracks_d
                all_finished.extend(finished_d)
                all_crossings.extend(crossings_d)   # ← key: keep every event
        except queue.Empty:
            pass

        if latest_frame is not None:
            self._last_frame  = latest_frame
            self._last_tracks = latest_tracks

            # Apply all crossing events collected across every drained frame
            if all_crossings:
                for ev in all_crossings:
                    if ev["dir"] == "IN":
                        self.cross_line.count_in  += 1
                    else:
                        self.cross_line.count_out += 1
                    self.cross_line.events.append(ev)
                    self._record_crossing(ev)           # ← save to results / Excel
                    self._cross_flash[ev["id"]] = {     # ← speed flash overlay
                        "speed": ev["speed"], "dir": ev["dir"],
                        "cx":    ev["cx"],    "cy":  ev["cy"],
                        "t0":    time.perf_counter(),
                    }
                self.lbl_cross.configure(
                    text=f"Chiziq: IN {self.cross_line.count_in}"
                         f" | OUT {self.cross_line.count_out}")

            self._update_fps()
            self._update_status()

        if self._last_frame is not None:
            self._render()
        self._schedule_poll()

    def _update_fps(self):
        self._ftimes.append(time.perf_counter())
        if len(self._ftimes) > 1:
            elapsed = self._ftimes[-1] - self._ftimes[0]
            fps = (len(self._ftimes)-1) / max(elapsed, 1e-6)
            self.lbl_fps.configure(text=f"{fps:.1f} FPS")
        # Stabilizer status
        self.lbl_stab.configure(text=f"Stab: {self.stabilizer.status_text}")
        self.ls_stab2.configure(text=f"Stab: {self.stabilizer.status_text}")

    def _step(self, n):
        if not self.cap: return
        cur   = int(self.cap.get(cv2.CAP_PROP_POS_FRAMES))
        total = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))
        self.cap.set(cv2.CAP_PROP_POS_FRAMES, max(0, min(total-1, cur+n)))
        ok, f = self.cap.read()
        if ok:
            self._last_frame = f
            self._render()
        pos = int(self.cap.get(cv2.CAP_PROP_POS_FRAMES))
        self._slider_busy = True
        try:
            self.slider.set(pos)
        finally:
            self._slider_busy = False
        self.lbl_time.configure(text=f"{pos / self.fps:.1f}s")

    def _seek(self, v):
        if not self.cap or self._slider_busy:
            return
        pos = int(float(v))
        self.cap.set(cv2.CAP_PROP_POS_FRAMES, pos)
        ok, f = self.cap.read()
        if ok:
            self._last_frame = f
            self._render()
        self.lbl_time.configure(text=f"{pos / self.fps:.1f}s")

    # ─── ROI ─────────────────────────────────────────────────────────────────
    def _roi_start(self):
        self._drawing_roi = True
        self._drawing_line = False
        self.roi = []; self._roi_tmp = None

    # ─── Counting line ────────────────────────────────────────────────────────
    def _line_start(self):
        """Enter line-draw mode: two left-clicks define the line."""
        self._drawing_line = True
        self._drawing_roi  = False
        self._line_p1      = None
        self._line_tmp     = None
        self.cross_line.clear()
        self.lbl_cross.configure(text="Chiziq: IN 0 | OUT 0")
        if self._last_frame is not None:
            self._render()

    def _canvas_click(self, e):
        if self._last_frame is None:
            return
        vx, vy = self._c2v(e.x, e.y)

        if self._drawing_line:
            if not hasattr(self, "_line_p1") or self._line_p1 is None:
                self._line_p1 = [vx, vy]
            else:
                self.cross_line.define(self._line_p1, [vx, vy])
                self._drawing_line = False
                self._line_p1      = None
                self._line_tmp     = None
                # Push updated line to Worker (if running)
                if self.worker:
                    self.worker.set_cross_line(self.cross_line)
                self._render()
            return

        if self._drawing_roi:
            self.roi.append([vx, vy])
            self._render()

    def _canvas_rclick(self, e):
        if self._drawing_line:
            self._drawing_line = False; self._line_p1 = None; self._line_tmp = None
            self._render(); return
        if self._drawing_roi and len(self.roi) >= 3:
            self._drawing_roi = False; self._roi_tmp = None
            fh, fw = self._last_frame.shape[:2]
            mask = np.zeros((fh, fw), np.uint8)
            cv2.fillPoly(mask, [np.array(self.roi, np.int32)], 255)
            self._roi_mask = mask
            if self.worker:
                self.worker.set_roi(mask)
            self._update_status(); self._render()

    def _canvas_motion(self, e):
        if self._last_frame is None:
            return
        vx, vy = self._c2v(e.x, e.y)
        if self._drawing_line:
            self._line_tmp = [vx, vy]
            self._render()
        elif self._drawing_roi:
            self._roi_tmp = [vx, vy]
            self._render()

    def _c2v(self, ex, ey):
        return (max(0, min(int((ex - self.ox) * self.sx), 9999)),
                max(0, min(int((ey - self.oy) * self.sy), 9999)))

    def _clear_all(self):
        self._stop_play()
        self.roi = []; self._roi_mask = None; self._drawing_roi = False
        self._drawing_line = False; self._line_tmp = None
        self.cross_line.clear()
        self.lbl_cross.configure(text="Chiziq: IN 0 | OUT 0")
        self.tracker = Tracker(self.fps, self.cam)
        self.stabilizer.reset()
        self.results = []; self.recorded_ids = set(); self._cross_flash = {}
        self.rl.delete(0, "end")
        Track._next = 1
        self._update_status()
        if self._last_frame is not None: self._render()

    # ─── Render ──────────────────────────────────────────────────────────────
    def _render(self):
        if self._last_frame is None: return
        d  = self._last_frame.copy()
        fh, fw = d.shape[:2]
        cw = max(self.canvas.winfo_width(),  100)
        ch = max(self.canvas.winfo_height(), 100)
        scale = min(cw / fw, ch / fh)
        nw, nh = int(fw * scale), int(fh * scale)
        self.sx, self.sy = fw / max(nw,1), fh / max(nh,1)
        self.ox, self.oy = (cw-nw)//2, (ch-nh)//2

        # ROI overlay
        if len(self.roi) >= 3:
            ov = d.copy()
            cv2.fillPoly(ov, [np.array(self.roi, np.int32)], (20, 80, 20))
            cv2.addWeighted(ov, 0.15, d, 0.85, 0, d)
            cv2.polylines(d, [np.array(self.roi, np.int32)], True, (50, 220, 50), 2)
        elif len(self.roi) > 0:
            for i, pt in enumerate(self.roi):
                cv2.circle(d, tuple(pt), 5, (50, 220, 50), -1)
                if i > 0:
                    cv2.line(d, tuple(self.roi[i-1]), tuple(pt), (50,220,50), 2)
            if self._roi_tmp and self._drawing_roi:
                cv2.line(d, tuple(self.roi[-1]), tuple(self._roi_tmp), (50,220,50), 1)

        # Counting line (draw before tracks so tracks appear on top)
        self.cross_line.draw(d, scale=1.0)
        # Line-draw preview
        if self._drawing_line and hasattr(self, "_line_p1") and self._line_p1 and self._line_tmp:
            cv2.line(d, tuple(self._line_p1), tuple(self._line_tmp), (0, 255, 255), 1, cv2.LINE_AA)
            cv2.circle(d, tuple(self._line_p1), 5, (0, 200, 255), -1)

        # Tracks (boxes + ID labels only — no speed text during normal tracking)
        draw_tracks(d, self._last_tracks, 1.0,
                    show_boxes=self.var_boxes.get(),
                    show_speed=self.var_speed_lbl.get())

        # Speed flash overlay — shown 2.5 s after each line crossing
        if self.var_speed_lbl.get() and self._cross_flash:
            draw_cross_flash(d, self._cross_flash, 1.0,
                             self.var_limit.get(), time.perf_counter())

        # Info overlay (top-left)
        active = len(self._last_tracks)
        t_sec  = self.cap.get(cv2.CAP_PROP_POS_MSEC) / 1000.0 if self.cap else 0
        info   = f"t={t_sec:.2f}s  |  {active} avto"
        cv2.rectangle(d, (4,4), (280,32), (0,0,0), -1)
        cv2.putText(d, info, (10,24), cv2.FONT_HERSHEY_SIMPLEX, 0.52, (200,200,200), 1)

        # Speed limit (top-right)
        lim_txt = f"Chegara: {self.var_limit.get():.0f} km/h"
        (tw,_),_ = cv2.getTextSize(lim_txt, cv2.FONT_HERSHEY_SIMPLEX, 0.48, 1)
        cv2.rectangle(d, (fw-tw-18,4), (fw-4,28), (0,0,0), -1)
        cv2.putText(d, lim_txt, (fw-tw-10,22),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.48, (100,200,255), 1)

        small = cv2.resize(d, (nw, nh), interpolation=cv2.INTER_LINEAR)
        small = cv2.cvtColor(small, cv2.COLOR_BGR2RGB)
        self._tk = ImageTk.PhotoImage(Image.fromarray(small))
        self.canvas.delete("all")
        self.canvas.create_image(self.ox, self.oy, anchor="nw", image=self._tk)

        # Time label
        if self.cap:
            total = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))
            pos   = int(self.cap.get(cv2.CAP_PROP_POS_FRAMES))
            self.lbl_time.configure(text=f"{t_sec:.1f}s / {total/self.fps:.1f}s")
            self._slider_busy = True
            try:
                self.slider.set(pos)
            finally:
                self._slider_busy = False

        # Header info
        speeds = [t.speed_kmh for t in self._last_tracks if t.speed_kmh > 1]
        if speeds:
            self.lbl_info.configure(
                text=f"{active} avto  |  avg {np.mean(speeds):.0f} km/h")
        else:
            self.lbl_info.configure(text=f"{active} avto" if active else "")

    # ─── Status ──────────────────────────────────────────────────────────────
    def _update_status(self):
        roi_ok = len(self.roi) >= 3
        self.ls_roi.configure(text=f"ROI: {'✓' if roi_ok else '✗'}",
                              fg="#34d399" if roi_ok else "#f87171")
        cam_txt = "Perspektiv ✓" if self.cam.enabled else "—"
        self.ls_cam.configure(text=f"Kamera: {cam_txt}",
                              fg="#34d399" if self.cam.enabled else "#fbbf24")
        if self.tracker:
            self.ls_trk.configure(
                text=f"Treklar: {len(self.tracker.tracks)} aktiv")

    # ─── Results recording ───────────────────────────────────────────────────
    def _record_crossing(self, ev: dict):
        """Record a single line-crossing event. Called once per unique track_id."""
        tid = ev["id"]
        if tid in self.recorded_ids:
            return
        self.recorded_ids.add(tid)
        entry = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "id":        tid,
            "direction": ev["dir"],
            "speed":     ev["speed"],   # 0.0 when no ROI is active
            "frame":     ev["frame"],
        }
        self.results.append(entry)
        spd_str = f"{ev['speed']:.1f}" if ev["speed"] > 0 else "—"
        tag = "→IN" if ev["dir"] == "IN" else "←OUT"
        self.rl.insert("end", f"ID:{tid} {tag}  {spd_str} km/h")
        self.rl.see("end")
        self._refresh_stats()

    def _refresh_stats(self):
        n = len(self.results)
        self.lbl_total.configure(text=f"Jami: {n} avto")
        if n:
            with_spd = [r["speed"] for r in self.results if r["speed"] > 0]
            avg  = np.mean(with_spd)  if with_spd else 0.0
            mx   = max(with_spd)      if with_spd else 0.0
            over = sum(1 for r in self.results
                       if r["speed"] > 0 and r["speed"] > self.var_limit.get())
            self.lbl_avg.configure(text=f"O'rtacha: {avg:.1f} km/h" if with_spd else "O'rtacha: —")
            self.lbl_max_s.configure(text=f"Maks: {mx:.1f} km/h"    if with_spd else "Maks: —")
            self.lbl_over.configure(text=f"Tezkor: {over}")

    # ─── Export ──────────────────────────────────────────────────────────────
    def _screenshot(self):
        if self._last_frame is None: return
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fp = os.path.join(self.ssdir, f"speed_{ts}.png")
        cv2.imwrite(fp, self._last_frame)
        messagebox.showinfo("Saqlandi", fp)

    def _export_excel(self):
        if not self.results:
            messagebox.showinfo("Info", "Hali chiziqdan o'tgan avtomobil yo'q."); return
        if not EXCEL_OK:
            messagebox.showerror("Xato", "openpyxl o'rnatilmagan.\npip install openpyxl")
            return
        save_dir = os.path.join(os.getcwd(), "excel_results")
        os.makedirs(save_dir, exist_ok=True)
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")],
            initialdir=save_dir,
            initialfile=f"speed_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        if not path: return
        self._write_excel(path)
        messagebox.showinfo("Tayyor!", f"Saqlandi:\n{path}")

    def _do_save_excel(self, auto=False):
        if not EXCEL_OK or not self.results: return
        save_dir = os.path.join(os.getcwd(), "excel_results")
        os.makedirs(save_dir, exist_ok=True)
        path = os.path.join(save_dir,
                            f"speed_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        self._write_excel(path)
        print(f"[Auto-save] {path}")

    def _write_excel(self, path: str):
        """Write one row per crossing event. Speed is present only when ROI is set."""
        wb  = Workbook()
        ws  = wb.active
        ws.title = "Chiziq Hodisalari"

        hf  = Font(name="Arial", size=11, bold=True, color="FFFFFFFF")
        hfl = PatternFill("solid", fgColor="FF1A5276")
        in_f  = Font(name="Arial", size=10, bold=True, color="FF007a33")
        out_f = Font(name="Arial", size=10, bold=True, color="FFCC0000")
        nf  = Font(name="Arial", size=10)
        bd  = Border(left=Side("thin"), right=Side("thin"),
                     top=Side("thin"), bottom=Side("thin"))
        ct  = Alignment(horizontal="center", vertical="center")

        # ── Title + meta ─────────────────────────────────────────────────────
        ws.merge_cells("A1:F1")
        ws["A1"] = "Speed Analyzer Pro — Chiziqdan O'tgan Avtomobillar"
        ws["A1"].font      = Font(name="Arial", size=14, bold=True, color="FFb45309")
        ws["A1"].alignment = ct
        ws.merge_cells("A2:F2")
        roi_str = "Ha" if self._roi_mask is not None else "Yo'q"
        ws["A2"] = (f"Video: {self.video_name}  |  FPS: {self.fps:.0f}  |"
                    f"  Chegara: {self.var_limit.get():.0f} km/h  |"
                    f"  ROI: {roi_str}  |"
                    f"  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        ws["A2"].font      = Font(name="Arial", size=9, italic=True, color="FF777777")
        ws["A2"].alignment = ct

        # ── Column headers ────────────────────────────────────────────────────
        headers = ["№", "Vaqt", "Track ID", "Yo'nalish", "Tezlik (km/h)", "Kadr №"]
        hr_row  = 4
        for c, h in enumerate(headers, 1):
            cl = ws.cell(hr_row, c, h)
            cl.font = hf; cl.fill = hfl; cl.alignment = ct; cl.border = bd

        # ── Data rows ─────────────────────────────────────────────────────────
        for i, r in enumerate(self.results):
            rr   = hr_row + 1 + i
            is_in = r["direction"] == "IN"
            df   = in_f if is_in else out_f
            spd  = r["speed"] if r["speed"] > 0 else None
            over = spd is not None and spd > self.var_limit.get()

            ws.cell(rr, 1, i + 1).font      = nf
            ws.cell(rr, 2, r["timestamp"]).font = nf
            ws.cell(rr, 3, r["id"]).font    = Font(name="Arial", size=10, bold=True)
            ws.cell(rr, 4, r["direction"]).font = df
            if spd is not None:
                ws.cell(rr, 5, spd).font         = out_f if over else in_f
                ws.cell(rr, 5).number_format     = "0.0"
            else:
                ws.cell(rr, 5, "—").font         = nf
            ws.cell(rr, 6, r["frame"]).font  = nf

            for c in range(1, 7):
                ws.cell(rr, c).border    = bd
                ws.cell(rr, c).alignment = ct
            if over:
                for c in range(1, 7):
                    ws.cell(rr, c).fill = PatternFill("solid", fgColor="FFFFF0F0")

        # ── Summary block ─────────────────────────────────────────────────────
        sr = hr_row + len(self.results) + 3
        ws.merge_cells(f"A{sr}:C{sr}")
        ws[f"A{sr}"] = "UMUMIY STATISTIKA"
        ws[f"A{sr}"].font = Font(name="Arial", size=12, bold=True, color="FF1A5276")

        with_spd = [r["speed"] for r in self.results if r["speed"] > 0]
        stats = [
            ("Jami o'tgan avtomobillar",  len(self.results)),
            ("Kiruvchi (IN)",              self.cross_line.count_in),
            ("Chiquvchi (OUT)",            self.cross_line.count_out),
            ("O'rtacha tezlik",
             f"{np.mean(with_spd):.1f} km/h" if with_spd else "— (ROI yo'q)"),
            ("Maksimal tezlik",
             f"{max(with_spd):.1f} km/h"     if with_spd else "— (ROI yo'q)"),
            ("Tezlik chegarasi",           f"{self.var_limit.get():.0f} km/h"),
            ("Chegaradan oshganlar",
             sum(1 for s in with_spd if s > self.var_limit.get())),
            ("Kamera modeli",              "Ha" if self.cam.enabled else "Yo'q"),
            ("ROI tezlik zonasi",          "Ha" if self._roi_mask is not None else "Yo'q"),
            ("Stabilizatsiya",             "Ha" if self.var_stab.get() else "Yo'q"),
        ]
        for j, (lbl, val) in enumerate(stats):
            ws.cell(sr + 1 + j, 1, lbl).font = Font(name="Arial", size=10, bold=True)
            ws.cell(sr + 1 + j, 2,
                    str(val) if not isinstance(val, (int, float)) else val).font = nf

        for i, w in enumerate([5, 20, 10, 12, 16, 10], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        wb.save(path)

    # ─── Run ─────────────────────────────────────────────────────────────────
    def run(self):
        self.root.mainloop()
        self._stop_play()
        if self.cap:
            self.cap.release()


# ═══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    root = tk.Tk()
    SpeedAnalyzerPro(root).run()
