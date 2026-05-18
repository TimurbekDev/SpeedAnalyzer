"""
Single-sheet row-based Excel exporter for traffic analysis.

One row per crossed vehicle. Totals row at bottom.
Supports real-time in-memory accumulation + on-demand export.
"""
from __future__ import annotations

from datetime import datetime
from typing import List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OK = True
except ImportError:
    _OK = False


# ── Style constants ────────────────────────────────────────────────────────────

_HDR_BG  = "0D1526"
_HDR_FG  = "00E5FF"
_ODD_BG  = "080D18"
_EVEN_BG = "0A1020"
_TOT_BG  = "0A0F1E"
_TOT_FG  = "7C3AED"
_RED_FG  = "FF3366"
_OK_FG   = "00DC82"
_MONO    = "Consolas"

# (header label, column width)
_COLS = [
    ("#",              6),
    ("Timestamp",     20),
    ("Vehicle ID",    12),
    ("Lane",           8),
    ("Direction",     12),
    ("Speed (km/h)",  14),
    ("Rear Vehicle",  14),
    ("Rear Dist (m)", 14),
    ("Violation",     12),
    ("Type",          12),
]


def _thin(color: str = "1E3A5F") -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _font(color: str = "C8D3E0", bold: bool = False, sz: int = 9) -> Font:
    return Font(name=_MONO, size=sz, color=color, bold=bold)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT   = Alignment(horizontal="left",   vertical="center")


class TrafficReportExporter:
    """
    Row-per-vehicle Excel exporter with in-memory accumulation.

    Call append() for each crossing event, then export() to write the file.
    reset() clears all accumulated rows (call on new session / clear-all).
    """

    def __init__(self):
        self._rows: List[tuple] = []
        self._seen: set          = set()

    @staticmethod
    def available() -> bool:
        return _OK

    def reset(self) -> None:
        self._rows.clear()
        self._seen.clear()

    def append(self,
               result:      dict,
               dist_result  = None,
               speed_limit: float = 60.0) -> None:
        """
        Store one crossing event row.

        result       — dict with keys: id, speed_kmh, cross_dir, timestamp, video
        dist_result  — CrossingDistanceResult or None
        speed_limit  — used to determine violation flag at record time
        """
        vid = result.get("id")
        if vid in self._seen:
            return
        self._seen.add(vid)

        spd     = result.get("speed_kmh", 0)
        over    = spd > speed_limit
        lane    = dist_result.lane_id    if dist_result else "—"
        rear_id = (dist_result.rear_id   if dist_result and dist_result.rear_id   is not None
                   else None)
        rear_m  = (dist_result.distance_m if dist_result and dist_result.distance_m is not None
                   else None)

        self._rows.append((
            len(self._rows) + 1,          # sequential row number
            result.get("timestamp", ""),
            vid,
            lane,
            result.get("cross_dir", "—"),
            round(spd, 1),
            rear_id,
            round(rear_m, 2) if rear_m is not None else None,
            over,
            "Crossline",
        ))

    def export(self, path: str, speed_limit: float = 60.0) -> None:
        if not _OK:
            raise RuntimeError("openpyxl not installed — cannot export.")
        if not self._rows:
            raise ValueError("No crossing events to export.")

        wb = Workbook()
        ws = wb.active
        ws.title = "Traffic Analysis"
        ws.sheet_view.showGridLines = False

        # ── Header ────────────────────────────────────────────────────────────
        for ci, (label, width) in enumerate(_COLS, 1):
            c = ws.cell(row=1, column=ci, value=label)
            c.font      = _font(_HDR_FG, bold=True, sz=10)
            c.fill      = _fill(_HDR_BG)
            c.alignment = _CENTER
            c.border    = _thin()
            ws.column_dimensions[get_column_letter(ci)].width = width

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 22

        # ── Data rows ─────────────────────────────────────────────────────────
        last_dr = 1
        for dr, row in enumerate(self._rows, 2):
            last_dr = dr
            row_bg = _ODD_BG if dr % 2 == 0 else _EVEN_BG
            num, ts, vid, lane, direction, spd, rear_id, rear_m, over, rtype = row
            spd_color = _RED_FG if over else _OK_FG

            vals = [
                num,
                ts,
                f"#{vid}",
                str(lane),
                direction,
                spd,
                f"#{rear_id}" if rear_id is not None else "—",
                f"{rear_m:.2f}" if rear_m is not None else "—",
                "YES" if over else "NO",
                rtype,
            ]
            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=dr, column=ci, value=val)
                c.font      = _font(spd_color if ci in (6, 9) else "C8D3E0", sz=9)
                c.fill      = _fill(row_bg)
                c.alignment = _CENTER
                c.border    = _thin()

        # ── Totals row (gap of one blank row) ─────────────────────────────────
        speeds = [r[5] for r in self._rows if r[5] > 0]
        n      = len(self._rows)
        viols  = sum(1 for r in self._rows if r[8])
        avg_s  = sum(speeds) / len(speeds) if speeds else 0.0
        max_s  = max(speeds)               if speeds else 0.0
        min_s  = min(speeds)               if speeds else 0.0

        tot_row = last_dr + 2
        tot_vals = [
            "TOTALS",
            f"Total Vehicles: {n}",
            "",
            "",
            "",
            f"Avg Speed: {avg_s:.1f}",
            f"Max Speed: {max_s:.1f}",
            f"Min Speed: {min_s:.1f}",
            f"Violations: {viols}",
            "",
        ]
        for ci, val in enumerate(tot_vals, 1):
            c = ws.cell(row=tot_row, column=ci, value=val)
            c.font      = _font(_TOT_FG, bold=True, sz=9)
            c.fill      = _fill(_TOT_BG)
            c.alignment = _LEFT if ci == 2 else _CENTER
            c.border    = _thin("2A4070")
        ws.row_dimensions[tot_row].height = 20

        # Remove default blank sheet if present
        if "Sheet" in wb.sheetnames and wb["Sheet"] is not ws:
            del wb["Sheet"]

        wb.save(path)


def make_filename() -> str:
    return f"traffic_analysis_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
