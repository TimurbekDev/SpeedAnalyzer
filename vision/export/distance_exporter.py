"""
DistanceExporter — one row per vehicle; writes/auto-saves .xlsx.

Schema:
    Track ID | Speed (km/h) | Distance to Next (m) | Next Vehicle ID | Timestamp

Each vehicle that passes through the lane ROI is logged exactly once with its
speed and the gap to the vehicle ahead of it.  History is kept in memory and the
file is rewritten after every record (auto-save), so the spreadsheet on disk
always reflects the full session even if the app is closed mid-run.
"""

from __future__ import annotations

from datetime import datetime
from typing import List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OK = True
except ImportError:
    _OK = False


_COLUMNS = [
    ("Track ID",             12),
    ("Speed (km/h)",         14),
    ("Distance to Next (m)", 20),
    ("Next Vehicle ID",      16),
    ("Timestamp",            14),
]


class DistanceExporter:
    def __init__(self):
        self.rows: List[tuple] = []
        self._auto_path: Optional[str] = None

    @staticmethod
    def available() -> bool:
        return _OK

    def reset(self) -> None:
        self.rows.clear()

    def set_auto_path(self, path: str) -> None:
        """File to (re)write automatically after each measurement."""
        self._auto_path = path

    # ── Records ───────────────────────────────────────────────────────────────

    def add(self, track_id: int, speed: float,
            distance_m, next_id, timestamp: str) -> tuple:
        """One vehicle's record. distance_m / next_id may be None (no vehicle
        ahead) and are written as a dash."""
        dist = round(float(distance_m), 1) if distance_m is not None else "—"
        nxt  = next_id if next_id is not None else "—"
        row = (track_id, round(float(speed), 1), dist, nxt, timestamp)
        self.rows.append(row)
        if self._auto_path:
            try:
                self.save(self._auto_path)
            except Exception as ex:               # never let export crash the loop
                print(f"[EXPORT] auto-save failed: {ex}")
        return row

    # ── Workbook ──────────────────────────────────────────────────────────────

    def save(self, path: str) -> None:
        if not _OK:
            raise RuntimeError("openpyxl not installed — cannot export.")
        wb = Workbook()
        ws = wb.active
        ws.title = "Distances"
        ws.sheet_view.showGridLines = False

        hdr_fill = PatternFill("solid", fgColor="0D1526")
        hdr_font = Font(name="Consolas", size=10, bold=True, color="00E5FF")
        cell_fnt = Font(name="Consolas", size=10, color="C8D3E0")
        center   = Alignment(horizontal="center", vertical="center")
        side     = Side(style="thin", color="1E3A5F")
        border   = Border(left=side, right=side, top=side, bottom=side)

        for ci, (label, width) in enumerate(_COLUMNS, 1):
            c = ws.cell(row=1, column=ci, value=label)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = center; c.border = border
            ws.column_dimensions[get_column_letter(ci)].width = width
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 22

        for ri, row in enumerate(self.rows, 2):
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.font = cell_fnt; c.alignment = center; c.border = border

        wb.save(path)


def make_filename(tag: str = "") -> str:
    """Timestamped .xlsx name, optionally tagged with the ROI/lane name so each
    region auto-saves to its own file."""
    stamp = f"{datetime.now():%Y%m%d_%H%M%S}"
    safe  = "".join(c for c in tag if c.isalnum() or c in "-_").strip("_")
    return (f"lane_distances_{safe}_{stamp}.xlsx" if safe
            else f"lane_distances_{stamp}.xlsx")
