"""Main application class — wires backend to modern UI."""

import cv2
import numpy as np
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from PIL import Image, ImageTk
import threading
import queue
import time
import os
from datetime import datetime
from collections import deque

from config.defaults import CONF_DEFAULT, CAP_Q_SIZE, RES_Q_SIZE
from core.entities.track import Track
from vision.detection.loader        import load_yolo
from vision.detection.bg_detector   import BGDetector
from vision.detection.enhancer      import Enhancer
from vision.tracking.stabilizer     import VideoStabilizer
from vision.tracking.tracker        import Tracker
from vision.geometry.camera_model   import CameraModel
from vision.geometry.lane_counter   import CrossLine
from vision.rendering.frame_renderer import draw_tracks, draw_cross_flash
from vision.pipeline.worker         import Worker

from ui.theme.dark          import Theme
from ui.widgets.components  import TopBar, SidebarNav, PlaybackBar, ToastManager
from ui.dialogs.calibrator  import CalibratorWindow
from ui.panels.dashboard    import DashboardPanel
from ui.panels.detection    import DetectionPanel
from ui.panels.analytics    import AnalyticsPanel
from ui.panels.settings_panel import SettingsPanel
from ui.panels.export_panel import ExportPanel

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False


class SpeedAnalyzerModern:
    def __init__(self, root: tk.Tk):
        self.root = root
        self._session_start = time.time()
        self._init_state()
        self._load_backend()
        self._build_ui()
        self._ui_loop()

    # ── State ─────────────────────────────────────────────────────────────────
    def _init_state(self):
        self.cap        = None
        self.fps        = 30.0
        self.video_name = ""
        self.playing    = False
        self.worker     = None
        self.cap_thread = None

        self.in_q:  queue.Queue = queue.Queue(maxsize=CAP_Q_SIZE)
        self.out_q: queue.Queue = queue.Queue(maxsize=RES_Q_SIZE)

        self._last_frame:  np.ndarray | None = None
        self._last_tracks: list = []
        self._ftimes = deque(maxlen=30)

        self.roi:          list               = []
        self._drawing_roi: bool               = False
        self._roi_tmp:     list | None        = None
        self._roi_mask:    np.ndarray | None  = None

        self.sx = self.sy = 1.0
        self.ox = self.oy = 0

        self.results:      list = []
        self.recorded_ids: set  = set()
        self.ssdir = "speed_screenshots"
        os.makedirs(self.ssdir, exist_ok=True)
        self._slider_busy = False

        self.cross_line    = CrossLine()
        self._drawing_line = False
        self._line_p1:  list | None = None
        self._line_tmp: list | None = None
        self._cross_flash: dict = {}

        self._stab_dx = self._stab_dy = 0.0
        self._violations = 0
        self._total      = 0

        self.var_limit     = tk.DoubleVar(value=60.0)
        self.var_stab      = tk.BooleanVar(value=True)
        self.var_stabview  = tk.BooleanVar(value=True)
        self.var_cam_on    = tk.BooleanVar(value=False)
        self.var_cam_H     = tk.DoubleVar(value=6.0)
        self.var_cam_T     = tk.DoubleVar(value=15.0)
        self.var_cam_P     = tk.DoubleVar(value=0.0)
        self.var_cam_F     = tk.DoubleVar(value=800.0)
        self.var_min_a     = tk.IntVar(value=1500)
        self.var_max_a     = tk.IntVar(value=150000)
        self.var_trail     = tk.BooleanVar(value=True)
        self.var_boxes     = tk.BooleanVar(value=True)
        self.var_speed_lbl = tk.BooleanVar(value=True)
        self.var_autosave  = tk.BooleanVar(value=True)
        self.var_speed_mul = tk.DoubleVar(value=1.0)
        self.var_speed_k   = tk.DoubleVar(value=1.0)
        self.var_conf      = tk.DoubleVar(value=CONF_DEFAULT)
        self.var_lane_hw   = tk.IntVar(value=CrossLine.DEFAULT_HW)

    def _load_backend(self):
        self.yolo_det, self.model_name = load_yolo()
        self.bg_det    = BGDetector()
        self.enhancer  = Enhancer()
        self.stabilizer = VideoStabilizer()
        self.cam       = CameraModel()
        self.tracker   = Tracker(30.0, self.cam)

    # ── UI build ──────────────────────────────────────────────────────────────
    def _build_ui(self):
        self.root.title("Speed Analyzer Pro")
        self.root.geometry("1480x900")
        self.root.minsize(1200, 700)
        self.root.configure(bg=Theme.BG)

        self.topbar = TopBar(self.root, model_name=self.model_name)
        self.topbar.pack(fill="x")
        tk.Frame(self.root, bg=Theme.DIVIDER, height=1).pack(fill="x")

        self.toast = ToastManager(self.root)

        middle = tk.Frame(self.root, bg=Theme.BG)
        middle.pack(fill="both", expand=True)

        self.sidebar = SidebarNav(middle, self._navigate)
        self.sidebar.pack(side="left", fill="y")
        tk.Frame(middle, bg=Theme.DIVIDER, width=1).pack(side="left", fill="y")

        self.content = tk.Frame(middle, bg=Theme.BG)
        self.content.pack(side="left", fill="both", expand=True)

        tk.Frame(self.root, bg=Theme.DIVIDER, height=1).pack(fill="x", side="bottom")
        self.playbar = PlaybackBar(self.root, self)
        self.playbar.pack(fill="x", side="bottom")

        self._panels: dict = {
            "dashboard": DashboardPanel(self.content, self),
            "detection": DetectionPanel(self.content, self),
            "analytics": AnalyticsPanel(self.content, self),
            "settings":  SettingsPanel(self.content, self),
            "export":    ExportPanel(self.content, self),
        }
        for panel in self._panels.values():
            panel.place(relx=0, rely=0, relwidth=1, relheight=1)

        self._navigate("dashboard")

        self.canvas = self._panels["dashboard"].video.canvas
        self.canvas.bind("<ButtonPress-1>",  self._canvas_click)
        self.canvas.bind("<Motion>",         self._canvas_motion)
        self.canvas.bind("<ButtonPress-3>",  self._canvas_rclick)
        self.canvas.bind("<Configure>",
                         lambda _: self._render() if self._last_frame is not None else None)

    def _navigate(self, key: str):
        self._active_panel = key
        for k, p in self._panels.items():
            (p.lift if k == key else p.lower)()
        if key == "analytics":
            self._panels["analytics"].refresh(
                self.results, self._session_start, self.var_limit.get())
        if key == "export":
            self._panels["export"].refresh(self.results, self.var_limit.get())

    # ── 60-fps UI refresh ─────────────────────────────────────────────────────
    def _ui_loop(self):
        if self._last_tracks:
            speeds = [t.speed_kmh for t in self._last_tracks if t.speed_kmh > 1]
            avg = np.mean(speeds) if speeds else 0
            self._panels["dashboard"].update_cards(
                self._total, avg, self._violations, self._current_fps())
            self.topbar.set_tracks(len(self._last_tracks))
        self.root.after(33, self._ui_loop)

    def _current_fps(self) -> float:
        if len(self._ftimes) < 2:
            return 0.0
        dts = [self._ftimes[i+1] - self._ftimes[i]
               for i in range(len(self._ftimes) - 1)
               if self._ftimes[i+1] > self._ftimes[i]]
        return (1.0 / (sum(dts) / len(dts))) if dts else 0.0

    # ── Video open ────────────────────────────────────────────────────────────
    def _open_video(self):
        path = filedialog.askopenfilename(
            filetypes=[("Video", "*.mp4 *.avi *.mov *.mkv *.ts"), ("All", "*.*")])
        if not path:
            return
        self._stop_play()
        if self.cap:
            self.cap.release()
        self.cap = cv2.VideoCapture(path)
        self.fps = self.cap.get(cv2.CAP_PROP_FPS) or 30.0
        self.video_name = os.path.basename(path)
        total = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))
        self.playbar.slider.configure(to=max(1, total - 1))
        self.playbar.set_time(0, total / self.fps)
        self.tracker   = Tracker(self.fps, self.cam)
        self.results   = []; self.recorded_ids = set(); self._cross_flash = {}
        self.roi       = []; self._roi_mask = None
        self._total    = 0;  self._violations = 0
        self._panels["dashboard"].chart.reset()
        ok, f = self.cap.read()
        if ok:
            self._last_frame  = f
            self._last_tracks = []
            self._panels["dashboard"].video.set_has_frame()
            self._render()
        self.root.title(f"Speed Analyzer Pro — {self.video_name}")
        self.topbar.set_status("IDLE")
        self.topbar.set_model(self.model_name or "BG MODE")
        self.toast.show(f"Opened: {self.video_name}", "info")

    # ── Playback ──────────────────────────────────────────────────────────────
    def _toggle_play(self):
        if not self.cap:
            self.toast.show("Open a video file first.", "warning"); return
        if self.playing:
            self._stop_play()
        else:
            self._start_play()

    def _start_play(self):
        self.playing = True
        self.playbar.set_playing(True)
        self.topbar.set_status("PLAYING")
        for q in (self.in_q, self.out_q):
            while not q.empty():
                try: q.get_nowait()
                except queue.Empty: break
        if self.tracker is None:
            self.tracker = Tracker(self.fps, self.cam)
        self.tracker.reset()
        self.stabilizer.reset()
        self.bg_det = BGDetector(self.var_min_a.get(), self.var_max_a.get())
        self.worker = Worker(self.yolo_det, self.bg_det, self.enhancer,
                             self.tracker, self.in_q, self.out_q,
                             cross_line=self.cross_line)
        if self._roi_mask is not None:
            self.worker.set_roi(self._roi_mask)
        self.worker.start()
        self.cap_thread = threading.Thread(target=self._cap_loop, daemon=True)
        self.cap_thread.start()
        self._schedule_poll()

    def _stop_play(self):
        if not self.playing:
            return
        self.playing = False
        self.playbar.set_playing(False)
        self.topbar.set_status("IDLE")
        try:
            self.in_q.put_nowait(None)
        except queue.Full:
            pass
        if self.worker:
            self.worker.stop()
            self.worker = None
        if self.var_autosave.get() and self.results:
            self._do_save_excel(auto=True)

    def _cap_loop(self):
        interval = 1.0 / self.fps
        t_next   = time.perf_counter() + interval
        while self.playing:
            ok, frame = self.cap.read()
            if not ok:
                try: self.in_q.put(None, timeout=1.0)
                except queue.Full: pass
                break
            stab_frame, dx, dy, _ = self.stabilizer.process(frame)
            self._stab_dx, self._stab_dy = dx, dy
            try:
                fid = int(self.cap.get(cv2.CAP_PROP_POS_FRAMES))
                self.in_q.put_nowait((fid, stab_frame))
            except queue.Full:
                pass
            sleep_t = t_next - time.perf_counter()
            if sleep_t > 0:
                time.sleep(sleep_t)
            t_next += interval

    def _schedule_poll(self):
        delay = max(8, int(1000.0 / self.fps))
        self.root.after(delay, self._poll)

    def _poll(self):
        if not self.playing:
            return
        latest_frame  = None
        latest_tracks = None
        all_finished:  list = []
        all_crossings: list = []

        try:
            while True:
                item = self.out_q.get_nowait()
                if item is None:
                    self._stop_play()
                    self.toast.show("Video finished.", "info")
                    return
                frame_d, tracks_d, finished_d, crossings_d = item
                latest_frame  = frame_d
                latest_tracks = tracks_d
                all_finished.extend(finished_d)
                all_crossings.extend(crossings_d)
        except queue.Empty:
            pass

        if latest_frame is not None:
            self._last_frame  = latest_frame
            self._last_tracks = latest_tracks

            if all_crossings:
                for ev in all_crossings:
                    if ev["dir"] == "IN":
                        self.cross_line.count_in  += 1
                    else:
                        self.cross_line.count_out += 1
                    self.cross_line.events.append(ev)
                    self._record_crossing(ev)
                    self._cross_flash[ev["id"]] = {
                        "speed": ev["speed"], "dir": ev["dir"],
                        "cx":    ev["cx"],    "cy":  ev["cy"],
                        "t0":    time.perf_counter(),
                    }
                self._panels["dashboard"].update_crossing(
                    self.cross_line.count_in, self.cross_line.count_out)

            if latest_tracks:
                spd = max((t.speed_kmh for t in latest_tracks), default=0)
                if spd > 1:
                    self._panels["dashboard"].chart.push(spd, self.var_limit.get())

            self._update_fps()
            self._panels["dashboard"].video.set_has_frame()

        if self._last_frame is not None:
            self._render()
            if self.cap:
                pos   = int(self.cap.get(cv2.CAP_PROP_POS_FRAMES))
                total = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))
                self._slider_busy = True
                try:
                    self.playbar.slider.set(pos)
                    self.playbar.set_time(pos / self.fps, total / self.fps)
                finally:
                    self._slider_busy = False

        self._schedule_poll()

    def _update_fps(self):
        self._ftimes.append(time.perf_counter())
        self.topbar.set_fps(self._current_fps())

    def _record_crossing(self, ev: dict):
        track_id = ev["id"]
        if track_id in self.recorded_ids:
            return
        self.recorded_ids.add(track_id)
        speed = ev["speed"]
        limit = self.var_limit.get()
        self._total += 1
        if speed > limit:
            self._violations += 1
            if speed > limit * 1.3:
                self.toast.show(f"Speed Alert: #{track_id}  {speed:.0f} km/h",
                                "error", duration=4000)
        record = {
            "id":        track_id,
            "speed_kmh": round(speed, 1),
            "cross_dir": ev["dir"],
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "video":     self.video_name,
        }
        self.results.append(record)
        self._panels["dashboard"].log_detection(track_id, speed, ev["dir"], limit)

    # ── Step / seek ───────────────────────────────────────────────────────────
    def _step(self, n: int):
        if not self.cap:
            return
        cur   = int(self.cap.get(cv2.CAP_PROP_POS_FRAMES))
        total = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))
        self.cap.set(cv2.CAP_PROP_POS_FRAMES, max(0, min(total - 1, cur + n)))
        ok, f = self.cap.read()
        if ok:
            self._last_frame = f
            self._panels["dashboard"].video.set_has_frame()
            self._render()
        pos = int(self.cap.get(cv2.CAP_PROP_POS_FRAMES))
        self._slider_busy = True
        try:
            self.playbar.slider.set(pos)
            self.playbar.set_time(pos / self.fps, total / self.fps)
        finally:
            self._slider_busy = False

    def _seek(self, v):
        if not self.cap or self._slider_busy:
            return
        pos = int(float(v))
        self.cap.set(cv2.CAP_PROP_POS_FRAMES, pos)
        ok, f = self.cap.read()
        if ok:
            self._last_frame = f
            self._panels["dashboard"].video.set_has_frame()
            self._render()
        if self.cap:
            total = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))
            self.playbar.set_time(pos / self.fps, total / self.fps)

    # ── Canvas / ROI / line ───────────────────────────────────────────────────
    def _roi_start(self):
        self._drawing_roi  = True
        self._drawing_line = False
        self.roi = []; self._roi_tmp = None
        self.toast.show("Left-click to add ROI points. Right-click to finish.", "info")

    def _line_start(self):
        self._drawing_line = True
        self._drawing_roi  = False
        self._line_p1      = None
        self._line_tmp     = None
        self.cross_line.clear()
        self.cross_line.set_width(self.var_lane_hw.get())
        self._panels["dashboard"].update_crossing(0, 0)
        if self._last_frame is not None:
            self._render()
        self.toast.show("Click two points to define the counting line.", "info")

    def _clear_line(self):
        self.cross_line.clear()
        self._drawing_line = False
        self._line_p1 = self._line_tmp = None
        self._panels["dashboard"].update_crossing(0, 0)
        if self._last_frame is not None:
            self._render()

    def _canvas_click(self, e):
        if self._last_frame is None:
            return
        vx, vy = self._c2v(e.x, e.y)
        if self._drawing_line:
            if self._line_p1 is None:
                self._line_p1 = [vx, vy]
            else:
                self.cross_line.define(self._line_p1, [vx, vy])
                self.cross_line.set_width(self.var_lane_hw.get())
                self._drawing_line = False
                self._line_p1      = None
                self._line_tmp     = None
                if self.worker:
                    self.worker.set_cross_line(self.cross_line)
                self._render()
                hw = self.var_lane_hw.get()
                self.toast.show(f"Counting line set  |  corridor ±{hw}px", "success")
            return
        if self._drawing_roi:
            self.roi.append([vx, vy])
            self._render()

    def _canvas_rclick(self, e):
        if self._drawing_line:
            self._drawing_line = False
            self._line_p1 = self._line_tmp = None
            self._render()
            return
        if self._drawing_roi and len(self.roi) >= 3:
            self._drawing_roi = False
            self._roi_tmp     = None
            fh, fw = self._last_frame.shape[:2]
            mask = np.zeros((fh, fw), np.uint8)
            cv2.fillPoly(mask, [np.array(self.roi, np.int32)], 255)
            self._roi_mask = mask
            if self.worker:
                self.worker.set_roi(mask)
            self._render()
            self.toast.show("ROI applied.", "success")

    def _canvas_motion(self, e):
        if self._last_frame is None:
            return
        vx, vy = self._c2v(e.x, e.y)
        if self._drawing_line:
            self._line_tmp = [vx, vy]
            self._render()
        elif self._drawing_roi:
            self._roi_tmp = [vx, vy]
            self._render()

    def _c2v(self, ex: int, ey: int):
        return (max(0, min(int((ex - self.ox) * self.sx), 9999)),
                max(0, min(int((ey - self.oy) * self.sy), 9999)))

    def _clear_all(self):
        self._stop_play()
        self.roi           = []; self._roi_mask = None
        self._drawing_roi  = False; self._drawing_line = False
        self._line_tmp     = None; self._line_p1 = None
        self.cross_line.clear()
        self._panels["dashboard"].update_crossing(0, 0)
        self.tracker       = Tracker(self.fps, self.cam)
        self.stabilizer.reset()
        self.results       = []; self.recorded_ids = set()
        self._cross_flash  = {}; self._total = 0; self._violations = 0
        Track._next        = 1
        self._panels["dashboard"].chart.reset()
        if self._last_frame is not None:
            self._render()
        self.toast.show("All data cleared.", "warning")

    # ── Render ────────────────────────────────────────────────────────────────
    def _render(self):
        if self._last_frame is None:
            return
        d  = self._last_frame.copy()
        fh, fw = d.shape[:2]
        cw = max(self.canvas.winfo_width(),  100)
        ch = max(self.canvas.winfo_height(), 100)
        scale = min(cw / fw, ch / fh)
        nw, nh = int(fw * scale), int(fh * scale)
        self.sx = fw / max(nw, 1)
        self.sy = fh / max(nh, 1)
        self.ox = (cw - nw) // 2
        self.oy = (ch - nh) // 2

        if len(self.roi) >= 3:
            ov = d.copy()
            cv2.fillPoly(ov, [np.array(self.roi, np.int32)], (20, 80, 20))
            cv2.addWeighted(ov, 0.15, d, 0.85, 0, d)
            cv2.polylines(d, [np.array(self.roi, np.int32)], True, (50, 220, 50), 2)
        elif self.roi:
            for i, pt in enumerate(self.roi):
                cv2.circle(d, tuple(pt), 5, (50, 220, 50), -1)
                if i > 0:
                    cv2.line(d, tuple(self.roi[i-1]), tuple(pt), (50, 220, 50), 2)
            if self._roi_tmp and self._drawing_roi:
                cv2.line(d, tuple(self.roi[-1]), tuple(self._roi_tmp),
                         (50, 220, 50), 1)

        self.cross_line.draw(d)
        if self._drawing_line and self._line_p1 and self._line_tmp:
            cv2.line(d, tuple(self._line_p1), tuple(self._line_tmp),
                     (0, 255, 255), 1, cv2.LINE_AA)
            cv2.circle(d, tuple(self._line_p1), 5, (0, 200, 255), -1)

        draw_tracks(d, self._last_tracks, 1.0,
                    show_boxes=self.var_boxes.get(),
                    show_speed=self.var_speed_lbl.get())

        if self.var_speed_lbl.get() and self._cross_flash:
            now = time.perf_counter()
            draw_cross_flash(d, self._cross_flash, 1.0,
                             speed_limit=self.var_limit.get(), now=now)
            self._cross_flash = {k: v for k, v in self._cross_flash.items()
                                 if now - v["t0"] < 2.5}

        stab_txt = (f"Stab dx={self._stab_dx:.1f} dy={self._stab_dy:.1f}"
                    if self.stabilizer.enabled else "Stab: OFF")
        cv2.putText(d, stab_txt, (8, 20),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.45, (80, 160, 200), 1, cv2.LINE_AA)

        if nw > 0 and nh > 0:
            disp = cv2.resize(d, (nw, nh), interpolation=cv2.INTER_LINEAR)
            img  = ImageTk.PhotoImage(
                Image.fromarray(cv2.cvtColor(disp, cv2.COLOR_BGR2RGB)))
            self.canvas.delete("frame")
            self.canvas.create_image(self.ox, self.oy, anchor="nw",
                                     image=img, tags="frame")
            self.canvas._img_ref = img

    # ── Camera model ──────────────────────────────────────────────────────────
    def _apply_camera(self):
        self.cam.speed_factor = max(0.01, self.var_speed_k.get())
        if self.var_cam_on.get():
            fw = int(self.cap.get(cv2.CAP_PROP_FRAME_WIDTH))  if self.cap else 1280
            fh = int(self.cap.get(cv2.CAP_PROP_FRAME_HEIGHT)) if self.cap else  720
            self.cam.setup(self.var_cam_H.get(), self.var_cam_T.get(),
                           self.var_cam_P.get(), self.var_cam_F.get(), fw, fh)
            self.toast.show("Camera model enabled.", "success")
        else:
            self.cam.enabled = False

    def _load_cam_json(self):
        path = filedialog.askopenfilename(filetypes=[("JSON", "*.json")])
        if not path:
            return
        try:
            self.cam.load_json(path)
            self.var_cam_H.set(self.cam.H)
            self.var_cam_T.set(self.cam.tilt)
            self.var_cam_P.set(self.cam.pan)
            self.var_cam_F.set(self.cam.f)
            self.var_cam_on.set(True)
            self.toast.show("Camera JSON loaded.", "success")
        except Exception as ex:
            messagebox.showerror("Error", str(ex))

    def _open_calibrator(self):
        CalibratorWindow(self.root, callback=self._on_calib_result,
                         frame_ref=self._last_frame)

    def _on_calib_result(self, params: dict):
        self.cam.H    = params.get("camera_height_m",  self.cam.H)
        self.cam.tilt = params.get("camera_pitch_deg", self.cam.tilt)
        self.cam.f    = params.get("focal_length_px",  self.cam.f)
        fw = params.get("frame_width",  1280)
        fh = params.get("frame_height",  720)
        self.cam.cx, self.cam.cy = fw / 2.0, fh / 2.0
        if "px_per_m" in params:
            self.cam.px_per_m = params["px_per_m"]
        self.cam.enabled = True
        self.var_cam_on.set(True)
        self.var_cam_H.set(self.cam.H)
        self.var_cam_T.set(self.cam.tilt)
        self.var_cam_F.set(self.cam.f)
        self.toast.show("Camera calibrated.", "success")

    def _load_model(self):
        path = filedialog.askopenfilename(
            filetypes=[("Model", "*.pt *.onnx *.engine"), ("All", "*.*")])
        if not path:
            return
        self._stop_play()
        det, name = load_yolo(path)
        if det:
            self.yolo_det   = det
            self.model_name = name
            self.topbar.set_model(name, ok=True)
            det_panel = self._panels["detection"]
            if hasattr(det_panel, "lbl_model"):
                det_panel.lbl_model.config(text=name, fg=Theme.CYAN)
            self.toast.show(f"Model loaded: {name}", "success")
        else:
            self.topbar.set_model("Load failed", ok=False)
            self.toast.show("Failed to load model.", "error")

    # ── Screenshot ────────────────────────────────────────────────────────────
    def _take_screenshot(self):
        if self._last_frame is None:
            self.toast.show("No frame to capture.", "warning"); return
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(self.ssdir, f"screenshot_{ts}.png")
        cv2.imwrite(path, self._last_frame)
        self.toast.show(f"Screenshot saved: {path}", "success")

    # ── Excel export ──────────────────────────────────────────────────────────
    def _export_excel(self):
        if not EXCEL_OK:
            messagebox.showerror("Error", "openpyxl not installed."); return
        if not self.results:
            self.toast.show("No results to export.", "warning"); return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"speed_report_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
        if path:
            self._do_save_excel(path=path)

    def _do_save_excel(self, path: str = None, auto: bool = False):
        if not EXCEL_OK or not self.results:
            return
        if path is None:
            ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = f"speed_report_{ts}.xlsx"
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Speed Report"

            hdr_fill = PatternFill("solid", fgColor="0D1526")
            hdr_font = Font(bold=True, color="00E5FF", name="Consolas", size=10)
            def_font = Font(name="Consolas", size=9)
            red_font = Font(name="Consolas", size=9, color="FF3366")
            center   = Alignment(horizontal="center")
            thin     = Side(style="thin", color="1E3A5F")
            border   = Border(left=thin, right=thin, top=thin, bottom=thin)

            headers = ["ID", "Speed (km/h)", "Direction",
                       "Timestamp", "Video", "Over Limit"]
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.fill = hdr_fill; c.font = hdr_font
                c.alignment = center; c.border = border

            limit = self.var_limit.get()
            for ri, r in enumerate(self.results, 2):
                spd  = r.get("speed_kmh", 0)
                over = "YES" if spd > limit else "NO"
                row_vals = [r.get("id"), spd,
                            r.get("cross_dir", ""), r.get("timestamp", ""),
                            r.get("video", ""), over]
                for ci, val in enumerate(row_vals, 1):
                    c = ws.cell(row=ri, column=ci, value=val)
                    c.font = red_font if over == "YES" else def_font
                    c.alignment = center; c.border = border

            for ci, w in enumerate([8, 14, 12, 22, 30, 12], 1):
                ws.column_dimensions[get_column_letter(ci)].width = w

            wb.save(path)
            if not auto:
                self.toast.show(f"Exported: {os.path.basename(path)}", "success")
        except Exception as ex:
            if not auto:
                messagebox.showerror("Export Error", str(ex))

    def mainloop(self):
        self.root.mainloop()
